"""
M_Crm tablosundan 10 ornek satir cekip Excel dosyasi olarak kaydeder.
"""
import sys
import os
import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from decouple import config

SQL_SERVER_CONFIG = {
    'server': config('SQL_SERVER', default='100.109.143.127'),
    'port': config('SQL_PORT', default='14330'),
    'database': config('SQL_DATABASE', default='DerinSISShow'),
    'username': config('SQL_USERNAME', default='sa2'),
    'password': config('SQL_PASSWORD', default='1478236950Mm..'),
}

def connect_sql():
    drivers = ['{ODBC Driver 17 for SQL Server}', '{SQL Server}']
    available = pyodbc.drivers()
    driver = None
    for d in available:
        for candidate in ['ODBC Driver 17 for SQL Server', 'SQL Server']:
            if candidate in d:
                driver = '{' + d + '}'
                break
        if driver:
            break
    if not driver:
        driver = '{SQL Server}'

    conn_str = (
        f"DRIVER={driver};"
        f"SERVER={SQL_SERVER_CONFIG['server']},{SQL_SERVER_CONFIG['port']};"
        f"DATABASE={SQL_SERVER_CONFIG['database']};"
        f"UID={SQL_SERVER_CONFIG['username']};"
        f"PWD={SQL_SERVER_CONFIG['password']};"
        f"TrustServerCertificate=yes;"
        f"Connection Timeout=30"
    )
    return pyodbc.connect(conn_str)

def main():
    print("SQL Server'a baglaniliyor...")
    conn = connect_sql()
    cursor = conn.cursor()

    # Tum sutunlari cek, 10 satir
    cursor.execute("SELECT TOP 10 * FROM M_Crm WITH (NOLOCK)")
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]

    print(f"Sutunlar ({len(columns)} adet): {columns}")
    print(f"Cekilen satir: {len(rows)}")

    # Excel olustur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "M_Crm Ornekler"

    # Baslik satiri stili
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Veri satirlari
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Kolon genislikleri
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Kaydet
    output_path = os.path.join(os.path.expanduser("~"), "Desktop", "M_Crm_Ornekler.xlsx")
    wb.save(output_path)
    print(f"\nExcel kaydedildi: {output_path}")
    conn.close()

if __name__ == "__main__":
    main()
