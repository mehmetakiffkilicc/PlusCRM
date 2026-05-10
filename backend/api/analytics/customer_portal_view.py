from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK, HTTP_404_NOT_FOUND, HTTP_401_UNAUTHORIZED
import sqlite3
import logging
import time
from datetime import datetime, timedelta
from .base import get_user_from_request, validate_data_source
from .. import db_engine

logger = logging.getLogger(__name__)

# Performans için kritik indeksler - uygulama başlangıcında bir kez çalışır
_indexes_created = False

def ensure_performance_indexes(cursor):
    """Kritik performans indekslerini oluştur"""
    global _indexes_created
    if _indexes_created:
        return

    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_satislar_musteri ON satislar(musteri_id)",
        "CREATE INDEX IF NOT EXISTS idx_satislar_musteri_tarih ON satislar(musteri_id, tarih DESC, saat DESC)",
        "CREATE INDEX IF NOT EXISTS idx_satislar_musteri_fis ON satislar(musteri_id, fis_no)",
        "CREATE INDEX IF NOT EXISTS idx_satislar_musteri_marka ON satislar(musteri_id, marka_id)",
        "CREATE INDEX IF NOT EXISTS idx_satislar_musteri_kat ON satislar(musteri_id, kategori_id)",
        "CREATE INDEX IF NOT EXISTS idx_musterionerileri_musteri ON MusteriOnerileri(musteri_id, oneri_skoru DESC)",
        "CREATE INDEX IF NOT EXISTS idx_musteridetayozet_musteri ON musteridetayozet(musteri_id)",
        "CREATE INDEX IF NOT EXISTS idx_musteriler_rfm ON musteriler(rfm_segment)",
        "CREATE INDEX IF NOT EXISTS idx_musteriler_rfm_id ON musteriler(rfm_segment, id DESC)",
        "CREATE INDEX IF NOT EXISTS idx_musterietiketler_musteri ON musterietiketler(musteri_id)",
        "CREATE INDEX IF NOT EXISTS idx_mdo_harcama ON musteridetayozet(toplam_harcama DESC)",
        "CREATE INDEX IF NOT EXISTS idx_mdo_ziyaret ON musteridetayozet(toplam_alisveris DESC)",
    ]

    is_pg = db_engine.DB_BACKEND == 'postgresql'
    for idx in indexes:
        try:
            cursor.execute(idx)
        except Exception as e:
            logger.warning(f"Index creation warning: {e}")
            if is_pg:
                try:
                    cursor.connection.rollback()
                except Exception:
                    pass

    _indexes_created = True
    logger.info("Performance indexes ensured")

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_list(request, data_source_id):
    """Müşteri listesi ve detaylı filtreleme (musteridetayozet JOIN'li)"""
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz erişim'}, status=HTTP_401_UNAUTHORIZED)
    if not validate_data_source(user, data_source_id):
        return Response({'error': 'Veri kaynağı bulunamadı veya erişim izni yok'}, status=404)

    # Filtreler (PascalCase ve snake_case uyumlu)
    search = request.GET.get('search', '').strip()
    selected_segments = request.GET.getlist('segments')
    selected_labels = request.GET.get('etiketler', '').strip()
    
    min_spend = request.GET.get('min_spend')
    max_spend = request.GET.get('max_spend')
    min_visits = request.GET.get('min_visits')
    max_visits = request.GET.get('max_visits')
    activity_status = request.GET.get('activity_status')
    trend = request.GET.get('trend')
    churn_risk = request.GET.get('churn_risk')
    
    region = request.GET.get('region')
    customer_type = request.GET.get('customer_type')
    approval_status = request.GET.get('approval_status')

    page = int(request.GET.get('page', 1))
    skip_count = request.GET.get('skip_count', 'false').lower() in ('true', '1', 'yes')
    page_size = min(int(request.GET.get('limit', 20)), 200)

    conn = db_engine.get_connection()
    cursor = db_engine.get_dict_cursor(conn)
    # ensure_performance_indexes(cursor) # Skip during normal requests to avoid hang
    ph = db_engine.ph()

    params = []
    where_clauses = ["1=1"]

    if search:
        if db_engine.DB_BACKEND == 'postgresql':
            where_clauses.append(f"(m.ad LIKE {ph} OR m.telefon LIKE {ph} OR CAST(m.id AS TEXT) LIKE {ph})")
        else:
            where_clauses.append(f"(m.ad LIKE {ph} OR m.telefon LIKE {ph} OR m.id LIKE {ph})")
        search_val = f"%{search}%"
        params.extend([search_val, search_val, search_val])

    # 'segments' (list) veya 'rfm_segment' (tekil) parametresini destekle
    if not selected_segments:
        single_seg = request.GET.get('rfm_segment', '').strip()
        if single_seg:
            selected_segments = [single_seg]
    if selected_segments:
        if len(selected_segments) == 1 and ',' in selected_segments[0]:
            selected_segments = selected_segments[0].split(',')
        seg_placeholders = ','.join([ph] * len(selected_segments))
        if db_engine.DB_BACKEND == 'postgresql':
            where_clauses.append(f"m.rfm_segment IN ({seg_placeholders})")
        else:
            where_clauses.append(f"m.rfm_segment IN ({seg_placeholders})")
        params.extend(selected_segments)

    if region:
        where_clauses.append(f"m.kayit_magazasi = {ph}")
        params.append(region)
    if customer_type:
        where_clauses.append(f"m.tip = {ph}")
        params.append(customer_type)
    if approval_status:
        where_clauses.append(f"m.onay_durumu = {ph}")
        params.append(approval_status)

    is_pg = db_engine.DB_BACKEND == 'postgresql'
    mdo_toplam_harcama = 'toplam_harcama' if is_pg else 'ToplamHarcama'
    mdo_toplam_alisveris = 'toplam_alisveris' if is_pg else 'ToplamAlisveris'
    mdo_aktivite_durumu = 'aktivite_durumu' if is_pg else 'AktiviteDurumu'
    mdo_trend = 'trend' if is_pg else 'Trend'
    mdo_churn_risk_skoru = 'churn_risk_skoru' if is_pg else 'ChurnRiskSkoru'
    mdo_musteri_id = 'musteri_id' if is_pg else 'MusteriID'
    mdo_son_alisveris_tarihi = 'son_alisveris_tarihi' if is_pg else 'SonAlisverisTarihi'
    mdo_ortalama_sepet = 'ortalama_sepet_tutari' if is_pg else 'OrtalamaSepetTutari'

    if min_spend:
        where_clauses.append(f"o.{mdo_toplam_harcama} >= {ph}")
        params.append(float(min_spend))
    if max_spend:
        where_clauses.append(f"o.{mdo_toplam_harcama} <= {ph}")
        params.append(float(max_spend))
    if min_visits:
        where_clauses.append(f"o.{mdo_toplam_alisveris} >= {ph}")
        params.append(int(min_visits))
    if max_visits:
        where_clauses.append(f"o.{mdo_toplam_alisveris} <= {ph}")
        params.append(int(max_visits))
    if activity_status:
        where_clauses.append(f"o.{mdo_aktivite_durumu} = {ph}")
        params.append(activity_status)
    if trend:
        where_clauses.append(f"o.{mdo_trend} = {ph}")
        params.append(trend)
    if churn_risk:
        if churn_risk == 'Yuksek':
            where_clauses.append(f"o.{mdo_churn_risk_skoru} >= 70")
        elif churn_risk == 'Orta':
            where_clauses.append(f"o.{mdo_churn_risk_skoru} BETWEEN 30 AND 69")
        elif churn_risk == 'Dusuk':
            where_clauses.append(f"o.{mdo_churn_risk_skoru} < 30")

    basket_segment = request.GET.get('basket_segment')
    if basket_segment:
        if basket_segment == 'kucuk':
            where_clauses.append(f"o.{mdo_ortalama_sepet} < 200")
        elif basket_segment == 'orta':
            where_clauses.append(f"o.{mdo_ortalama_sepet} BETWEEN 200 AND 999.99")
        elif basket_segment == 'buyuk':
            where_clauses.append(f"o.{mdo_ortalama_sepet} BETWEEN 1000 AND 2999.99")
        elif basket_segment == 'mega':
            where_clauses.append(f"o.{mdo_ortalama_sepet} >= 3000")

    label_join_sql = ""
    if selected_labels:
        label_join_sql = "JOIN musterietiketler el ON m.id = el.musteri_id"
        labels = [l.strip() for l in selected_labels.split(',') if l.strip()]
        for lbl in labels:
            if lbl.endswith('_skoru'):
                where_clauses.append(f"el.{lbl} > 0.5")
            else:
                where_clauses.append(f"el.{lbl} = TRUE")

    where_sql = " AND ".join(where_clauses)
    
    # Sorting
    sort_by = request.GET.get('sort_by', 'id')
    sort_order = request.GET.get('sort_order', 'desc').upper()
    sort_mapping = {
        'id': 'm.id',
        'ad': 'm.ad',
        'total_spend': f'o.{mdo_toplam_harcama}',
        'total_visits': f'o.{mdo_toplam_alisveris}',
        'last_shopping': f'o.{mdo_son_alisveris_tarihi}',
        'churn_risk': f'o.{mdo_churn_risk_skoru}'
    }
    order_column = sort_mapping.get(sort_by, 'm.id')
    if is_pg:
        order_sql = f"{order_column} {sort_order} NULLS LAST"
    else:
        order_sql = f"CASE WHEN {order_column} IS NULL THEN 1 ELSE 0 END, {order_column} {sort_order}"

    # beklenenmusteriler tablosu var mı kontrol et
    beklenen_join = ""
    beklenen_select = ""
    try:
        cursor.execute("SELECT 1 FROM information_schema.tables WHERE table_name = 'beklenenmusteriler' LIMIT 1")
        if cursor.fetchone():
            beklenen_join = " LEFT JOIN beklenenmusteriler bk ON m.id = bk.musteri_id"
            beklenen_select = ", CASE WHEN bk.musteri_id IS NOT NULL THEN TRUE ELSE FALSE END AS bekleniyor, bk.gecikme_gun AS beklenen_gecikme_gun"
    except Exception:
        pass

    count_query = f"SELECT COUNT(DISTINCT m.id) as cnt FROM musteriler m LEFT JOIN musteridetayozet o ON m.id = o.{mdo_musteri_id} {label_join_sql} WHERE {where_sql}"

    try:
        if page == 1 and not skip_count:
            cursor.execute(count_query, params)
            total_count = db_engine.val(cursor.fetchone(), 'cnt')
        else:
            total_count = -1

        offset = (page - 1) * page_size
        query = f"""
            SELECT m.id, m.ad, m.telefon, m.tip, m.rfm_segment, m.kayit_tarihi,
                   COALESCE(o.{mdo_toplam_harcama}, 0) as total_spend,
                   COALESCE(o.{mdo_toplam_alisveris}, 0) as total_visits,
                   COALESCE(o.{mdo_ortalama_sepet}, 0) as avg_basket,
                   o.{mdo_aktivite_durumu} as activity_status, o.{mdo_trend} as trend,
                   COALESCE(o.{mdo_churn_risk_skoru}, 0) as churn_risk,
                   o.{mdo_son_alisveris_tarihi} as last_shopping_date,
                   el.sadik_musteri, el.gizli_risk, el.kaybedilmemesi_gereken, el.soguyan_musteri
                   {beklenen_select}
            FROM musteriler m
            LEFT JOIN musteridetayozet o ON m.id = o.{mdo_musteri_id}
            LEFT JOIN musterietiketler el ON m.id = el.musteri_id
            {beklenen_join}
            WHERE {where_sql}
            ORDER BY {order_sql}
            LIMIT {ph} OFFSET {ph}
        """
        cursor.execute(query, params + [page_size, offset])
        customers = [dict(row) for row in cursor.fetchall()]
        
        return Response({
            'customers': customers,
            'total': total_count,
            'page': page,
            'page_size': page_size
        })
    except Exception as e:
        logger.error(f"Customer list error: {e}")
        return Response({'error': str(e)}, status=500)
    finally:
        db_engine.release_connection(conn)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_detail(request, data_source_id, customer_id):
    """Müşteri derinlemesine analiz dashboard verisi - musteridetayozet destekli"""
    user = get_user_from_request(request)
    if not user: return Response({'error': 'Yetkisiz'}, status=401)
    if not validate_data_source(user, data_source_id):
        return Response({'error': 'Veri kaynağı bulunamadı veya erişim izni yok'}, status=404)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()
        
        # 1. Fiş Kalem Detayı (Bir fiş tıklandığında)
        fis_no_param = request.GET.get('fis_no', '').strip()
        if fis_no_param:
            cursor.execute(f"""
                SELECT s.tarih, s.saat, s.fis_no, s.urun_id, u.ad as urun_ad, m.ad as magaza_ad, s.miktar, s.tutar
                FROM satislar s
                JOIN urunler u ON s.urun_id = u.id
                JOIN magazalar m ON s.magaza_id = m.id
                WHERE s.musteri_id = {ph} AND s.fis_no = {ph}
                ORDER BY s.tutar DESC
            """, (customer_id, fis_no_param))
            items = [dict(row) for row in cursor.fetchall()]
            return Response({'history': items})

        # 2. Fiş Listesi (Sayfalı)
        if request.GET.get('mode') == 'fis_listesi':
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 100))
            offset = (page - 1) * page_size
            total_fis = -1
            if page == 1:
                cursor.execute(f"SELECT COUNT(DISTINCT fis_no) as total FROM satislar WHERE musteri_id = {ph}", (customer_id,))
                total_fis = cursor.fetchone()['total']
            cursor.execute(f"""
                SELECT s.tarih, s.saat, s.fis_no, m.ad as magaza_ad,
                       COUNT(DISTINCT s.fis_no) as kalem_sayisi, SUM(s.tutar) as toplam_tutar
                FROM satislar s JOIN magazalar m ON s.magaza_id = m.id
                WHERE s.musteri_id = {ph} 
                GROUP BY s.tarih, s.saat, s.fis_no, m.ad
                ORDER BY s.tarih DESC, s.saat DESC, s.fis_no DESC
                LIMIT {ph} OFFSET {ph}
            """, (customer_id, page_size, offset))
            fis_listesi = [dict(row) for row in cursor.fetchall()]
            return Response({'fis_listesi': fis_listesi, 'total_fis': total_fis, 'has_more': len(fis_listesi) == page_size})

        # 3. Ürün Analizi (tekrar alım + segment karşılaştırma)
        if request.GET.get('mode') == 'urun_analizi':
            # 3a. Tekrar satın alma listesi (fiş bazlı)
            cursor.execute(f"""
                SELECT u.ad as urun_ad, u.id as urun_id,
                       COUNT(DISTINCT s.fis_no) as fis_count,
                       CAST(SUM(s.miktar) AS FLOAT) as total_qty,
                       CAST(SUM(s.tutar) AS FLOAT) as total_spend,
                       MAX(s.tarih) as last_buy
                FROM satislar s JOIN urunler u ON s.urun_id = u.id
                WHERE s.musteri_id = {ph}
                GROUP BY u.id, u.ad
                ORDER BY fis_count DESC, total_spend DESC
                LIMIT 20
            """, (customer_id,))
            tekrar_alim = [dict(r) for r in cursor.fetchall()]

            # 3b. Müşterinin aldığı kategoriler
            cursor.execute(f"""
                SELECT DISTINCT k.ana as kategori_ad
                FROM satislar s JOIN kategoriler k ON s.kategori_id = k.id
                WHERE s.musteri_id = {ph} AND k.ana IS NOT NULL
            """, (customer_id,))
            musteri_kategorileri = {r['kategori_ad'] for r in cursor.fetchall()}

            # 3c. Segmentin popüler kategorileri (segmenturuntercihleri tablosundan)
            rfm_seg = None
            cursor.execute(f"SELECT rfm_segment FROM musteriler WHERE id = {ph}", (customer_id,))
            seg_row = cursor.fetchone()
            if seg_row:
                rfm_seg = db_engine.val(seg_row, 'rfm_segment')

            segment_kategoriler = []
            if rfm_seg:
                try:
                    cursor.execute(f"""
                        SELECT kategori_ad, AVG(penetrasyon) as ort_penetrasyon
                        FROM segmenturuntercihleri
                        WHERE rfm_segment = {ph} AND kategori_ad IS NOT NULL
                        GROUP BY kategori_ad
                        ORDER BY ort_penetrasyon DESC
                        LIMIT 15
                    """, (rfm_seg,))
                    rows = cursor.fetchall()
                    for r in rows:
                        kat_ad = db_engine.val(r, 'kategori_ad')
                        penetrasyon = float(db_engine.val(r, 'ort_penetrasyon') or 0)
                        segment_kategoriler.append({
                            'kategori_ad': kat_ad,
                            'penetrasyon': round(penetrasyon * 100, 1),
                            'musteri_aldi': kat_ad in musteri_kategorileri
                        })
                except Exception as e:
                    logger.warning(f"segmenturuntercihleri sorgu hatası: {e}")

            # 3d. Kategori derinlik analizi
            cursor.execute(f"""
                SELECT k.ana as kategori_ad,
                       COUNT(DISTINCT s.urun_id) as urun_cesidi,
                       COUNT(DISTINCT s.fis_no) as fis_count,
                       CAST(SUM(s.tutar) AS FLOAT) as toplam_harcama
                FROM satislar s JOIN kategoriler k ON s.kategori_id = k.id
                WHERE s.musteri_id = {ph} AND k.ana IS NOT NULL
                GROUP BY k.ana
                ORDER BY toplam_harcama DESC
            """, (customer_id,))
            kategori_derinlik = [dict(r) for r in cursor.fetchall()]

            # 3e. Hane etiketi bazlı öne çıkan ürünler
            # Bebek/çocuk skoru yüksek müşteriler için ilgili kategorideki ürünleri öne çıkar
            hane_vurgular = []
            try:
                cursor.execute(f"""
                    SELECT hane_bebek_skoru, hane_cocuklu_skoru
                    FROM musterietiketler WHERE musteri_id = {ph}
                """, (customer_id,))
                etiket_row = cursor.fetchone()
                if etiket_row:
                    bebek_skoru = float(db_engine.val(etiket_row, 'hane_bebek_skoru') or 0)
                    cocuklu_skoru = float(db_engine.val(etiket_row, 'hane_cocuklu_skoru') or 0)

                    vurgu_kategoriler = []
                    if bebek_skoru >= 0.3:
                        vurgu_kategoriler.append(('bebek', bebek_skoru))
                    if cocuklu_skoru >= 0.3:
                        vurgu_kategoriler.append(('cocuk', cocuklu_skoru))

                    for (etiket, skor) in vurgu_kategoriler:
                        if etiket == 'bebek':
                            where_clause = "(k.ana LIKE '%bebek%' OR k.alt1 LIKE '%bebek%')"
                            label = 'Bebek Ürünleri'
                        else:
                            where_clause = "(k.alt1 LIKE '%oyuncak%' OR k.alt1 LIKE '%çocuk%' OR k.alt1 LIKE '%kırtasiye%')"
                            label = 'Çocuk Ürünleri'

                        cursor.execute(f"""
                            SELECT u.ad as urun_ad, u.id as urun_id,
                                   COUNT(DISTINCT s.fis_no) as fis_count,
                                   CAST(SUM(s.miktar) AS FLOAT) as total_qty,
                                   CAST(SUM(s.tutar) AS FLOAT) as total_spend
                            FROM satislar s
                            JOIN urunler u ON s.urun_id = u.id
                            JOIN kategoriler k ON u.kategori_id = k.id
                            WHERE s.musteri_id = {ph} AND {where_clause}
                            GROUP BY u.id, u.ad
                            ORDER BY fis_count DESC, total_spend DESC
                            LIMIT 10
                        """, (customer_id,))
                        urunler = [dict(r) for r in cursor.fetchall()]
                        if urunler:
                            hane_vurgular.append({
                                'etiket': etiket,
                                'label': label,
                                'skor': round(skor, 2),
                                'urunler': urunler
                            })
            except Exception as e:
                logger.warning(f"Hane vurgu ürünleri sorgu hatası: {e}")

            return Response({
                'tekrar_alim': tekrar_alim,
                'segment_kategoriler': segment_kategoriler,
                'kategori_derinlik': kategori_derinlik,
                'musteri_kategori_sayisi': len(musteri_kategorileri),
                'hane_vurgular': hane_vurgular,
            })

        # 4. ANA DASHBOARD VERİSİ
        t0 = time.time()
        is_pg = db_engine.DB_BACKEND == 'postgresql'
        mdo_musteri_id = 'musteri_id' if is_pg else 'MusteriID'

        # 3.1. Temel Bilgiler + Özet Tablo
        cursor.execute(f"""
            SELECT m.*, o.*
            FROM musteriler m
            LEFT JOIN musteridetayozet o ON m.id = o.{mdo_musteri_id}
            WHERE m.id = {ph}
        """, (customer_id,))
        customer_data = cursor.fetchone()
        if not customer_data: return Response({'error': 'Müşteri bulunamadı'}, status=404)
        info = dict(customer_data)

        # 3.2. KPI'lar (HIZLI PATH + Self-Healing Fallback)
        kpis = {
            'total_visits': int(db_engine.val(info, 'toplam_alisveris') or 0),
            'total_spend': float(db_engine.val(info, 'toplam_harcama') or 0),
            'total_units': float(db_engine.val(info, 'toplam_miktar_calculated') or 0),
            'last_shopping_date': db_engine.val(info, 'son_alisveris_tarihi'),
            'first_shopping_date': db_engine.val(info, 'ilk_alisveris_tarihi'),
            'avg_basket': float(db_engine.val(info, 'ortalama_sepet_tutari') or 0),
            'ltv': float(db_engine.val(info, 'lifetime_value_tahmini') or 0),
            'churn_risk': float(db_engine.val(info, 'churn_risk_skoru') or 0),
            'trend': db_engine.val(info, 'trend') or 'Stabil',
            'activity_status': db_engine.val(info, 'aktivite_durumu') or 'Pasif',
            'fav_category': db_engine.val(info, 'favori_kategori') or '-',
            'fav_brand': db_engine.val(info, 'favori_marka') or '-',
            'fav_store': db_engine.val(info, 'favori_magaza') or '-',
            'fav_product': db_engine.val(info, 'favori_urun') or '-',
            'preferred_hour': str(db_engine.val(info, 'tercih_edilen_saat') or '00'),
            'preferred_day': db_engine.val(info, 'gun_tercihi') or '-',
            'spend_30d': float(db_engine.val(info, 'son_30_gun_harcama') or 0),
            'spend_90d': float(db_engine.val(info, 'son_90_gun_harcama') or 0),
            'loyalty_ratio': float(db_engine.val(info, 'marka_sadakati') or 0),
            'morning_count': int(db_engine.val(info, 'saat_sabah') or 0),
            'noon_count': int(db_engine.val(info, 'saat_ogle') or 0),
            'evening_count': int(db_engine.val(info, 'saat_aksam') or 0),
            'night_count': int(db_engine.val(info, 'saat_gece') or 0),
            # Ek alanlar
            'visits_30d': int(db_engine.val(info, 'son_30_gun_alisveris') or 0),
            'visits_90d': int(db_engine.val(info, 'son_90_gun_alisveris') or 0),
            'customer_age_days': int(db_engine.val(info, 'musteri_yasi_gun') or 0),
            'basket_diversity': float(db_engine.val(info, 'sepet_cesitlendirme') or 0),
        }

        # DÖNEMSEL KPI'lar (Tarih filtresi varsa anlık hesapla)
        if start_date or end_date:
            try:
                period_where = [f"musteri_id = {ph}"]
                period_params = [customer_id]
                if start_date: 
                    period_where.append(f"tarih >= {ph}")
                    period_params.append(start_date)
                if end_date:
                    period_where.append(f"tarih <= {ph}")
                    period_params.append(end_date)
                
                period_where_clause = " AND ".join(period_where)
                # PostgreSQL/SQLite ph conversion handled by db_engine if needed, but here we use satislar usually
                period_query = f"""
                    SELECT SUM(tutar) as total_spend, COUNT(DISTINCT fis_no) as total_visits, 
                           SUM(miktar) as total_units, MIN(tarih) as first_date, MAX(tarih) as last_date
                    FROM satislar 
                    WHERE {period_where_clause}
                """
                cursor.execute(db_engine.adapt_query(period_query), period_params)
                period_data = cursor.fetchone()
                
                if period_data:
                    kpis['total_spend'] = float(db_engine.val(period_data, 'total_spend') or 0)
                    kpis['total_visits'] = int(db_engine.val(period_data, 'total_visits') or 0)
                    kpis['total_units'] = float(db_engine.val(period_data, 'total_units') or 0)
                    kpis['avg_basket'] = kpis['total_spend'] / kpis['total_visits'] if kpis['total_visits'] > 0 else 0
                    # if last_date is in period, use it, otherwise keep overall
                    p_last = db_engine.val(period_data, 'last_date')
                    if p_last: kpis['last_shopping_date'] = p_last
                    p_first = db_engine.val(period_data, 'first_date')
                    if p_first: kpis['first_shopping_date'] = p_first
            except Exception as e:
                logger.error(f"Periodical KPI calculation error: {e}")

        # SELF-HEALING: Eğer özet tablo boş veya 0 ise, satislar tablosundan anlık hesapla
        if kpis['total_spend'] == 0:
            try:
                cursor.execute(f"SELECT SUM(tutar) as total, COUNT(DISTINCT fis_no) as visits, MAX(tarih) as last_date FROM satislar WHERE musteri_id = {ph}", (customer_id,))
                real_data = cursor.fetchone()
                if real_data and db_engine.val(real_data, 'total'):
                    kpis['total_spend'] = float(db_engine.val(real_data, 'total') or 0)
                    kpis['total_visits'] = int(db_engine.val(real_data, 'visits') or 0)
                    kpis['last_shopping_date'] = db_engine.val(real_data, 'last_date')
                    if kpis['total_visits'] > 0:
                        kpis['avg_basket'] = kpis['total_spend'] / kpis['total_visits']
                    kpis['activity_status'] = 'Aktif'
            except Exception as e:
                logger.warning(f"Self-healing KPI calculation failed: {e}")

        # 3.3. Favori Markalar (CTE ile hızlı)
        fav_brands = []
        if is_pg:
            cursor.execute(f"""
                WITH top_m AS (
                    SELECT marka_id, SUM(tutar) as revenue, SUM(miktar) as qty
                    FROM satislar WHERE musteri_id = {ph}
                    GROUP BY marka_id ORDER BY revenue DESC LIMIT 5
                )
                SELECT tm.marka_id as id, mr.ad as name,
                       CAST(tm.revenue AS FLOAT) as revenue,
                       CAST(tm.qty AS FLOAT) as qty
                FROM top_m tm JOIN markalar mr ON tm.marka_id = mr.id
                ORDER BY tm.revenue DESC
            """, (customer_id,))
        else:
            cursor.execute(f"""
                SELECT s.marka_id as id, m.ad as name,
                       CAST(SUM(s.tutar) AS FLOAT) as revenue,
                       CAST(SUM(s.miktar) AS FLOAT) as qty
                FROM satislar s JOIN markalar m ON s.marka_id = m.id
                WHERE s.musteri_id = {ph}
                GROUP BY s.marka_id, m.ad ORDER BY revenue DESC LIMIT 5
            """, (customer_id,))
        fav_brands = [dict(r) for r in cursor.fetchall()]
        
        # Marka bazlı en çok alınan ürünler
        for brand in fav_brands:
            cursor.execute(f"""
                SELECT s.urun_id as id, u.ad as name, CAST(SUM(s.miktar) AS FLOAT) as total_qty, CAST(SUM(s.tutar) AS FLOAT) as total_revenue
                FROM satislar s
                JOIN urunler u ON s.urun_id = u.id
                WHERE s.musteri_id = {ph} AND s.marka_id = {ph}
                GROUP BY s.urun_id, u.ad
                ORDER BY total_revenue DESC
                LIMIT 3
            """, (customer_id, brand['id']))
            brand['top_products'] = [dict(r) for r in cursor.fetchall()]

        # 3.4. Favori Kategoriler
        cursor.execute(f"""
            SELECT k.id as id, TRIM(k.ana) as name, CAST(SUM(s.tutar) AS FLOAT) as revenue
            FROM satislar s JOIN kategoriler k ON s.kategori_id = k.id
            WHERE s.musteri_id = {ph}
            GROUP BY k.id, TRIM(k.ana) ORDER BY revenue DESC
            LIMIT 30
        """, (customer_id,))
        category_distribution = [dict(r) for r in cursor.fetchall()]
        fav_categories = category_distribution[:5]
        
        # Kategori bazlı en çok alınan ürünler
        for cat in fav_categories:
            cursor.execute(f"""
                SELECT s.urun_id as id, u.ad as name, CAST(SUM(s.miktar) AS FLOAT) as total_qty, CAST(SUM(s.tutar) AS FLOAT) as total_revenue
                FROM satislar s
                JOIN urunler u ON s.urun_id = u.id
                WHERE s.musteri_id = {ph} AND s.kategori_id = {ph}
                GROUP BY s.urun_id, u.ad
                ORDER BY total_revenue DESC
                LIMIT 3
            """, (customer_id, cat['id']))
            cat['top_products'] = [dict(r) for r in cursor.fetchall()]

        # 3.5. Son 50 alışveriş kalemi
        cursor.execute(f"""
            SELECT s.tarih, s.saat, s.fis_no, s.urun_id, u.ad as urun_ad, mg.ad as magaza_ad, s.miktar, s.tutar, k.ana as kategori_ad
            FROM satislar s 
            JOIN urunler u ON s.urun_id = u.id 
            JOIN magazalar mg ON s.magaza_id = mg.id
            JOIN kategoriler k ON s.kategori_id = k.id
            WHERE s.musteri_id = {ph} ORDER BY s.tarih DESC, s.saat DESC LIMIT 50
        """, (customer_id,))
        history = [dict(row) for row in cursor.fetchall()]

        # 3.6. Saat dağılımı
        cursor.execute(f"""
            SELECT CAST(saat AS INTEGER) as hour, COUNT(DISTINCT musteri_id) as count
            FROM satislar WHERE musteri_id = {ph} AND saat IS NOT NULL
            GROUP BY hour ORDER BY hour
        """, (customer_id,))
        time_dist = [dict(row) for row in cursor.fetchall()]

        # 3.7. Gün dağılımı (fiş bazlı - kalem değil)
        cursor.execute(f"""
            SELECT {db_engine.strftime_expr('%w', 'tarih')} as day_num,
                   COUNT(DISTINCT fis_no) as count,
                   SUM(tutar) as total_spend
            FROM satislar WHERE musteri_id = {ph} GROUP BY day_num ORDER BY day_num
        """, (customer_id,))
        day_distribution = [dict(row) for row in cursor.fetchall()]

        # 3.8. Harcama Trendi (Son 90 gün)
        ninety_days_ago = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        cursor.execute(f"""
            SELECT {db_engine.strftime_expr('%Y-%W', 'tarih')} as week, MIN(tarih) as week_start,
                   SUM(tutar) as total_spend, COUNT(DISTINCT fis_no) as visit_count
            FROM satislar WHERE musteri_id = {ph} AND tarih >= {ph}
            GROUP BY week ORDER BY week
        """, (customer_id, ninety_days_ago))
        spending_trend = [dict(row) for row in cursor.fetchall()]

        # 3.9. RFM Skorları
        # musteridetayozet'te r_score/f_score/m_score, musteriler'de rfm_r_score/rfm_f_score/rfm_m_score
        rfm_scores = {
            'recency': int(db_engine.val(info, 'r_score') or db_engine.val(info, 'rfm_r_score') or 0),
            'frequency': int(db_engine.val(info, 'f_score') or db_engine.val(info, 'rfm_f_score') or 0),
            'monetary': int(db_engine.val(info, 'm_score') or db_engine.val(info, 'rfm_m_score') or 0),
            'segment': db_engine.val(info, 'rfm_segment') or 'Belirsiz'
        }

        # 3.9b. Ek Metrikler (Yıllık Harcama & Sıklık)
        one_year_ago = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        cursor.execute(f"SELECT SUM(tutar) as total FROM satislar WHERE musteri_id = {ph} AND tarih >= {ph}", (customer_id, one_year_ago))
        kpis['spend_365d'] = float(db_engine.val(cursor.fetchone(), 'total') or 0)
        
        # Alışveriş Sıklığı (Son 1 yıldaki ortalama gün aralığı)
        if kpis['total_visits'] > 1:
            # İlk ve son alışveriş arası gün farkı / (ziyaret - 1)
            try:
                cursor.execute(f"SELECT MIN(tarih) as first, MAX(tarih) as last FROM satislar WHERE musteri_id = {ph}", (customer_id,))
                dates = cursor.fetchone()
                if dates['first'] and dates['last']:
                    d1 = datetime.strptime(str(dates['first']), '%Y-%m-%d')
                    d2 = datetime.strptime(str(dates['last']), '%Y-%m-%d')
                    days_diff = (d2 - d1).days
                    if days_diff > 0:
                        kpis['avg_visit_interval'] = round(days_diff / (kpis['total_visits'] - 1), 1)
                    else:
                        kpis['avg_visit_interval'] = 0
                else:
                    kpis['avg_visit_interval'] = 0
            except:
                kpis['avg_visit_interval'] = 0
        else:
            kpis['avg_visit_interval'] = 0

        # 3.10. Davranis Etiketleri (TAM LİSTE - Segmentation sayfamızla senkron)
        labels = {}
        churn_skoru = None
        try:
            cursor.execute(f"SELECT * FROM musterietiketler WHERE musteri_id = {ph}", (customer_id,))
            label_row = cursor.fetchone()
            if label_row:
                ld = dict(label_row)
                churn_skoru = ld.get('churn_skoru')
                # Grupları eksiksiz yükle (Hane skorları dahil)
                groups = {
                    'ziyaret': ['sabah_alisveriscisi', 'aksam_alisveriscisi', 'gece_alisveriscisi', 'hafta_sonu_alisveriscisi', 'hafta_ici_alisveriscisi', 'aylik_duzenli_alici', 'maas_gunu_alisveriscisi', 'gunluk_ugrayan', 'seyrek_alisverisci', 'cok_magazali_musteri'],
                    'sepet': ['buyuk_sepet_alisveriscisi', 'kucuk_sepet_alisveriscisi', 'premium_harcayici', 'ekonomik_harcayici', 'b2b_mahalle_esnafi', 'stokcu_alici', 'tekli_urun_alisveriscisi', 'mixed_sepet_alisveriscisi'],
                    'fiyat': ['indirim_avcisi', 'promosyon_bagimli', 'fiyat_hassas', 'fiyata_duyarsiz', 'coklu_alim_firsatcisi', 'enflasyon_stokcusu', 'kampanya_tepkisi_dusuk'],
                    'taze_gida': ['kasap_odakli', 'manav_odakli', 'firinci_odakli', 'sarkuteri_odakli', 'sut_urunleri_odakli', 'bakliyat_odakli', 'baharat_odakli', 'sadece_taze_gidaci', 'yoresel_urun_meraklisi', 'taze_gida_kacinani', 'taze_gida_terk_eden'],
                    'urun_odak': ['atistirmalik_odakli', 'icecek_tutkunuodakli', 'kahvaltilik_odakli', 'kisisel_bakim_odakli', 'temizlik_odakli', 'bebek_urunleri_odakli', 'ev_tekstili_odakli', 'organik_saglikli_odakli', 'dondurulmus_odakli'],
                    'kategori': ['saglikli_yasam_egilimli', 'hazir_tuketim_egilimli', 'protein_odakli', 'kafein_yogun_tuketici', 'atistirmalik_tuketicisi', 'temizlik_hijyen_odakli', 'kisisel_bakim_tutkunu', 'misafir_sofrasi_kurucusu'],
                    'kanal': ['winback_adayi', 'reaktivasyon_potansiyeli', 'yeniden_kazanilmis', 'kampanya_duyarli', 'kampanyasiz_sadik'],
                    'odeme': ['yemek_karti_kullanicisi', 'ay_sonu_yemek_karti_harcayicisi', 'fatura_musterisi'],
                    'sadakat': ['sadik_musteri', 'soguyan_musteri', 'kaybedilme_riski_yuksek', 'tamamen_kaybedilmis', 'yeniden_kazanilmis_saglik', 'gidip_gelen_musteri', 'sepeti_daralan', 'kategori_terk_eden', 'marji_dusuran', 'gizli_risk', 'kaybedilmemesi_gereken'],
                    'hane': ['hane_bekar_skoru', 'hane_cift_skoru', 'hane_aile_skoru', 'hane_cocuklu_skoru', 'hane_bebek_skoru', 'hane_yasli_skoru', 'hane_evcil_hayvan_skoru', 'hane_araba_skoru', 'hane_toplu_alim_skoru']
                }
                labels = { g: { k: ld.get(k) for k in cols } for g, cols in groups.items() }
        except: pass

        # 3.11. Karşılaştırmalı KPI'lar
        cursor.execute("SELECT AVG(toplam_harcama) as avg_h, AVG(toplam_alisveris) as avg_a, AVG(ortalama_sepet_tutari) as avg_basket FROM musteridetayozet")
        averages = cursor.fetchone() or {'avg_h': 0, 'avg_a': 0, 'avg_basket': 0}
        kpis['spend_vs_avg'] = round((kpis['total_spend'] / (averages['avg_h'] or 1) - 1) * 100, 1)
        kpis['visits_vs_avg'] = round((kpis['total_visits'] / (averages['avg_a'] or 1) - 1) * 100, 1)

        # 3.12. Segment Benchmark (segmenturuntercihleri veya segment ortalaması)
        segment_benchmark = {'avg_basket': 0, 'avg_spend': 0, 'avg_visits': 0, 'customer_count': 0}
        try:
            rfm_seg_val = rfm_scores['segment']
            cursor.execute(f"""
                SELECT AVG(o.ortalama_sepet_tutari) as avg_basket,
                       AVG(o.toplam_harcama) as avg_spend,
                       AVG(o.toplam_alisveris) as avg_visits,
                        COUNT(DISTINCT el.musteri_id) as customer_count
                FROM musteridetayozet o
                JOIN musteriler m ON m.id = o.{mdo_musteri_id}
                WHERE m.rfm_segment = {ph}
            """, (rfm_seg_val,))
            bench = cursor.fetchone()
            if bench:
                segment_benchmark = {
                    'avg_basket': round(float(db_engine.val(bench, 'avg_basket') or 0), 2),
                    'avg_spend': round(float(db_engine.val(bench, 'avg_spend') or 0), 2),
                    'avg_visits': round(float(db_engine.val(bench, 'avg_visits') or 0), 1),
                    'customer_count': int(db_engine.val(bench, 'customer_count') or 0),
                }
        except Exception as e:
            logger.warning(f"Segment benchmark hatası: {e}")

        # 3.13. Fiyat Esnekliği (musterifiyatfeatures)
        fiyat_ozeti = None
        try:
            cursor.execute(f"""
                SELECT indirim_oran_yuzde, ort_indirim_yuzde, toplam_indirim_tutari, toplam_brut_tutar, toplam_satis_satir
                FROM musterifiyatfeatures WHERE musteri_id = {ph}
            """, (customer_id,))
            f_row = cursor.fetchone()
            if f_row:
                indirim_oran = float(db_engine.val(f_row, 'indirim_oran_yuzde') or 0)
                ort_indirim = float(db_engine.val(f_row, 'ort_indirim_yuzde') or 0)
                toplam_indirim = float(db_engine.val(f_row, 'toplam_indirim_tutari') or 0)
                toplam_brut = float(db_engine.val(f_row, 'toplam_brut_tutar') or 0)
                # İndirim hassasiyeti seviyesi: yüksek >%50, orta %20-50, düşük <%20
                if indirim_oran >= 50:
                    hassasiyet = 'Yüksek'
                    onerilen_indirim = f'%{max(10, round(ort_indirim * 0.8))}-{round(ort_indirim)}'
                elif indirim_oran >= 20:
                    hassasiyet = 'Orta'
                    onerilen_indirim = f'%{round(ort_indirim * 0.7)}-{round(ort_indirim * 0.9)}'
                else:
                    hassasiyet = 'Düşük'
                    onerilen_indirim = '%5-10'
                fiyat_ozeti = {
                    'indirim_oran_yuzde': round(indirim_oran, 1),
                    'ort_indirim_yuzde': round(ort_indirim, 1),
                    'toplam_indirim_tutari': round(toplam_indirim, 2),
                    'toplam_brut_tutar': round(toplam_brut, 2),
                    'hassasiyet_seviye': hassasiyet,
                    'onerilen_indirim_araligi': onerilen_indirim,
                }
        except Exception as e:
            logger.warning(f"Fiyat esnekliği hatası: {e}")

        # 3.14. Dönem Karşılaştırması (musteridonem_karsilastirma)
        donem_karsilastirma = None
        try:
            cursor.execute(f"""
                SELECT harcama_3ay, harcama_onceki3ay, harcama_degisim_3ay_yuzde,
                       ziyaret_3ay, ziyaret_onceki3ay, ziyaret_degisim_3ay_yuzde,
                       harcama_6ay, harcama_onceki6ay, harcama_degisim_6ay_yuzde,
                       ziyaret_6ay, ziyaret_onceki6ay, ziyaret_degisim_6ay_yuzde,
                       terk_edilen_kategori
                FROM musteridonem_karsilastirma WHERE musteri_id = {ph}
            """, (customer_id,))
            d_row = cursor.fetchone()
            if d_row:
                donem_karsilastirma = {
                    'harcama_3ay': round(float(db_engine.val(d_row, 'harcama_3ay') or 0), 2),
                    'harcama_onceki3ay': round(float(db_engine.val(d_row, 'harcama_onceki3ay') or 0), 2),
                    'harcama_degisim_3ay': round(float(db_engine.val(d_row, 'harcama_degisim_3ay_yuzde') or 0), 1),
                    'ziyaret_3ay': int(db_engine.val(d_row, 'ziyaret_3ay') or 0),
                    'ziyaret_onceki3ay': int(db_engine.val(d_row, 'ziyaret_onceki3ay') or 0),
                    'ziyaret_degisim_3ay': round(float(db_engine.val(d_row, 'ziyaret_degisim_3ay_yuzde') or 0), 1),
                    'harcama_degisim_6ay': round(float(db_engine.val(d_row, 'harcama_degisim_6ay_yuzde') or 0), 1),
                    'ziyaret_degisim_6ay': round(float(db_engine.val(d_row, 'ziyaret_degisim_6ay_yuzde') or 0), 1),
                    'terk_edilen_kategori': int(db_engine.val(d_row, 'terk_edilen_kategori') or 0),
                }
        except Exception as e:
            logger.warning(f"Dönem karşılaştırma hatası: {e}")

        # 3.14b. Kategori dağılımı (musterikategoridagilimi)
        kategori_dagilimi = []
        try:
            cursor.execute(f"""
                SELECT ana_kategori,
                       SUM(fis_sayisi) as fis_sayisi,
                       SUM(toplam_harcama) as toplam_harcama,
                       COUNT(DISTINCT alt_kategori) as alt_kategori_sayisi
                FROM musterikategoridagilimi
                WHERE musteri_id = {ph}
                GROUP BY ana_kategori
                ORDER BY toplam_harcama DESC
            """, (customer_id,))
            rows = cursor.fetchall()
            toplam_kat_harcama = sum(float(db_engine.val(r, 'toplam_harcama', 0)) for r in rows)
            for r in rows:
                harcama = round(float(db_engine.val(r, 'toplam_harcama', 0)), 2)
                kategori_dagilimi.append({
                    'ana_kategori': db_engine.val(r, 'ana_kategori', ''),
                    'fis_sayisi': int(db_engine.val(r, 'fis_sayisi', 0)),
                    'toplam_harcama': harcama,
                    'alt_kategori_sayisi': int(db_engine.val(r, 'alt_kategori_sayisi', 0)),
                    'oran': round(harcama / toplam_kat_harcama * 100, 1) if toplam_kat_harcama > 0 else 0,
                })
        except Exception as e:
            logger.warning(f"Kategori dağılımı hatası: {e}")

        # 3.15. Cross-sell kategorileri (grupbirliktelikleri)
        cross_sell_cats = []
        try:
            cursor.execute(f"""
                SELECT k2.alt2 as kategori_ad, k2.alt1 as alt_kategori, k2.ana as ana_kategori,
                       MAX(CAST(gb.lift AS FLOAT)) as lift,
                       MAX(CAST(gb.confidence AS FLOAT)) as confidence
                FROM grupbirliktelikleri gb
                JOIN kategoriler k1 ON gb.kategori_id_1 = k1.id
                JOIN kategoriler k2 ON gb.kategori_id_2 = k2.id
                WHERE gb.kategori_id_1 IN (
                    SELECT DISTINCT kategori_id FROM satislar WHERE musteri_id = {ph}
                )
                AND gb.kategori_id_2 NOT IN (
                    SELECT DISTINCT kategori_id FROM satislar WHERE musteri_id = {ph}
                )
                AND k2.alt2 IS NOT NULL
                GROUP BY k2.alt2, k2.alt1, k2.ana
                ORDER BY lift DESC
                LIMIT 5
            """, (customer_id, customer_id))
            cross_sell_cats = [dict(r) for r in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"Cross-sell hatası: {e}")

        logger.info(f"[PERF] Total detail time: {time.time()-t0:.3f}s for id={customer_id}")
        return Response({
            'info': info,
            'kpis': kpis,
            'fav_brands': fav_brands,
            'fav_categories': fav_categories,
            'history': history,
            'time_dist': time_dist,
            'day_distribution': day_distribution,
            'spending_trend': spending_trend,
            'rfm_scores': rfm_scores,
            'category_distribution': category_distribution,
            'labels': labels,
            'churn_skoru': churn_skoru,
            'segment_benchmark': segment_benchmark,
            'cross_sell_cats': cross_sell_cats,
            'fiyat_ozeti': fiyat_ozeti,
            'donem_karsilastirma': donem_karsilastirma,
            'kategori_dagilimi': kategori_dagilimi,
        })

    except Exception as e:
        logger.error(f"Customer detail error: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn: db_engine.release_connection(conn)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def musteri_birlestir(request, data_source_id):
    """İki müşteriyi birleştirir: Kaynak müşterinin tüm satışlarını hedef müşteriye taşır."""
    user = get_user_from_request(request)
    if not user: return Response({'error': 'Yetkisiz'}, status=401)
    
    source_id = request.data.get('source_id')
    target_id = request.data.get('target_id')
    
    if not source_id or not target_id:
        return Response({'error': 'Kaynak ve hedef müşteri ID gereklidir'}, status=400)
    
    if source_id == target_id:
        return Response({'error': 'Kaynak ve hedef aynı olamaz'}, status=400)

    conn = db_engine.get_connection()
    try:
        cursor = conn.cursor()
        ph = db_engine.ph()
        
        # 1. Satışları taşı
        cursor.execute(f"UPDATE satislar SET musteri_id = {ph} WHERE musteri_id = {ph}", (target_id, source_id))
        rows_moved = cursor.rowcount
        
        # 2. Kaynak müşteriyi sil (Opsiyonel: Kayıtlı müşteri tablosundan da çıkar)
        cursor.execute(f"DELETE FROM musteriler WHERE id = {ph}", (source_id,))
        cursor.execute(f"DELETE FROM musteridetayozet WHERE musteri_id = {ph}", (source_id,))
        cursor.execute(f"DELETE FROM musterietiketler WHERE musteri_id = {ph}", (source_id,))
        
        conn.commit()
        
        logger.info(f"Customer Merge: {source_id} merged into {target_id}. {rows_moved} sales moved.")
        return Response({
            'success': True, 
            'message': f'{source_id} ID\'li müşterinin {rows_moved} satışı {target_id} ID\'li müşteriye başarıyla aktarıldı.'
        })
    except Exception as e:
        if conn: conn.rollback()
        logger.error(f"Merge error: {e}")
        return Response({'error': str(e)}, status=500)
    finally:
        db_engine.release_connection(conn)

_etiket_ozeti_cache = {}
_etiket_ozeti_cache_timeout = 300  # 5 dakika

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_musteri_etiket_ozeti(request, data_source_id):
    """Etiket özeti (Segmentasyon sayfası için) - Tam Veri Yapısı Restorasyonu"""
    user = get_user_from_request(request)
    if not user: return Response({'error': 'Yetkisiz'}, status=401)

    customer_type = request.GET.get('customer_type') or request.GET.get('customerType')
    approval_status = request.GET.get('approval_status') or request.GET.get('approvalStatus')
    region = request.GET.get('region')
    has_filter = bool(customer_type or approval_status or region)

    # Cache: filtre yoksa cache kullan
    cache_key = f"etiket_ozeti_{data_source_id}"
    if not has_filter:
        # 1. Bellek içi cache (Hızlı)
        cached = _etiket_ozeti_cache.get(cache_key)
        if cached and (time.time() - cached['ts']) < _etiket_ozeti_cache_timeout:
            return Response(cached['data'])
        
        # 2. Veritabanı bazlı önbellek (Sync Worker tarafından doldurulur)
        try:
            conn = db_engine.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM syncmeta WHERE key = 'etiket_ozeti_global'")
            row = cursor.fetchone()
            if row:
                import json
                val = row[0] if isinstance(row[0], str) else row['value']
                data = json.loads(val)
                # Bellek içi cache'e de yaz
                _etiket_ozeti_cache[cache_key] = {'data': data, 'ts': time.time()}
                return Response(data)
        except Exception as e:
            logger.warning(f"Syncmeta cache read error: {e}")
        finally:
            if 'conn' in locals() and conn: db_engine.release_connection(conn)

    # Kategori ve Etiket Eşleşmeleri (Frontend: Segmentation.tsx ile tam uyumlu)
    LABEL_GROUPS = {
        'ziyaret': ['sabah_alisveriscisi', 'aksam_alisveriscisi', 'gece_alisveriscisi', 'hafta_sonu_alisveriscisi', 'hafta_ici_alisveriscisi', 'aylik_duzenli_alici', 'maas_gunu_alisveriscisi', 'gunluk_ugrayan', 'seyrek_alisverisci', 'cok_magazali_musteri'],
        'sepet': ['buyuk_sepet_alisveriscisi', 'kucuk_sepet_alisveriscisi', 'premium_harcayici', 'ekonomik_harcayici', 'b2b_mahalle_esnafi', 'stokcu_alici', 'tekli_urun_alisveriscisi', 'mixed_sepet_alisveriscisi'],
        'fiyat': ['indirim_avcisi', 'promosyon_bagimli', 'fiyat_hassas', 'fiyata_duyarsiz', 'coklu_alim_firsatcisi', 'enflasyon_stokcusu', 'kampanya_tepkisi_dusuk'],
        'taze_gida': ['kasap_odakli', 'manav_odakli', 'firinci_odakli', 'sarkuteri_odakli', 'sut_urunleri_odakli', 'bakliyat_odakli', 'baharat_odakli', 'sadece_taze_gidaci', 'yoresel_urun_meraklisi', 'taze_gida_kacinani', 'taze_gida_terk_eden'],
        'urun_odak': ['atistirmalik_odakli', 'icecek_tutkunuodakli', 'kahvaltilik_odakli', 'kisisel_bakim_odakli', 'temizlik_odakli', 'bebek_urunleri_odakli', 'ev_tekstili_odakli', 'organik_saglikli_odakli', 'dondurulmus_odakli'],
        'kategori': ['saglikli_yasam_egilimli', 'hazir_tuketim_egilimli', 'protein_odakli', 'kafein_yogun_tuketici', 'atistirmalik_tuketicisi', 'temizlik_hijyen_odakli', 'kisisel_bakim_tutkunu', 'misafir_sofrasi_kurucusu'],
        'kanal': ['winback_adayi', 'reaktivasyon_potansiyeli', 'yeniden_kazanilmis', 'kampanya_duyarli', 'kampanyasiz_sadik'],
        'odeme': ['yemek_karti_kullanicisi', 'ay_sonu_yemek_karti_harcayicisi', 'fatura_musterisi'],
        'sadakat': ['sadik_musteri', 'soguyan_musteri', 'kaybedilme_riski_yuksek', 'tamamen_kaybedilmis', 'yeniden_kazanilmis_saglik', 'gidip_gelen_musteri', 'sepeti_daralan', 'kategori_terk_eden', 'marji_dusuran', 'gizli_risk', 'kaybedilmemesi_gereken'],
        'hane': ['hane_bekar_skoru', 'hane_cift_skoru', 'hane_aile_skoru', 'hane_cocuklu_skoru', 'hane_bebek_skoru', 'hane_yasli_skoru', 'hane_evcil_hayvan_skoru', 'hane_araba_skoru', 'hane_toplu_alim_skoru']
    }

    SCORE_COLUMNS = {col for col in LABEL_GROUPS['hane']}
    all_labels = [label for group in LABEL_GROUPS.values() for label in group]
    
    conn = db_engine.get_connection()
    cursor = db_engine.get_dict_cursor(conn)
    try:
        ph = db_engine.ph()

        # Müşteri filtresi için JOIN koşulu oluştur
        if has_filter:
            join_parts = []
            join_params = []
            if customer_type:
                join_parts.append(f"m.tip = {ph}")
                join_params.append(customer_type)
            if approval_status:
                join_parts.append(f"m.onay_durumu = {ph}")
                join_params.append(approval_status)
            if region:
                join_parts.append(f"m.kayit_magazasi IN (SELECT id::text FROM magazalar WHERE {db_engine.bolge_expr()} = {ph})")
                join_params.append(region)
            join_where = " AND ".join(join_parts)
            filter_join = f"JOIN musteriler m ON me.musteri_id = m.id WHERE {join_where}"
        else:
            filter_join = ""
            join_params = []

        # 1. Toplam Müşteri
        cursor.execute(f"SELECT COUNT(DISTINCT musteri_id) as toplam FROM musterietiketler me {filter_join}", join_params)
        total = db_engine.val(cursor.fetchone(), 'toplam') or 1

        # 2. Tüm Etiket Sayıları (Tek bir SQL taraması ile)
        filter_parts = []
        for col in all_labels:
            if col in SCORE_COLUMNS:
                filter_parts.append(f"COUNT(DISTINCT CASE WHEN me.{col} >= 0.4 THEN me.{col} END) as {col}")
            else:
                filter_parts.append(f"COUNT(DISTINCT CASE WHEN me.{col} = TRUE THEN me.musteri_id END) as {col}")

        sql = f"SELECT {', '.join(filter_parts)} FROM musterietiketler me {filter_join}"
        cursor.execute(sql, join_params)
        counts = cursor.fetchone()

        # 3. Trend verisi: önceki snapshot ile karşılaştır
        trend_map = {}
        trend_tarihi = None
        try:
            cursor.execute("""
                SELECT etiket_kolon, sayi FROM etiket_snapshot
                WHERE tarih = (SELECT MAX(tarih) FROM etiket_snapshot WHERE tarih < CURRENT_DATE)
            """)
            prev_rows = cursor.fetchall()
            if prev_rows:
                cursor.execute("SELECT MAX(tarih) FROM etiket_snapshot WHERE tarih < CURRENT_DATE")
                trend_tarihi = str(cursor.fetchone()[0]) if cursor.fetchone is not None else None
                # tekrar çek çünkü fetchone tükendi
                cursor.execute("SELECT MAX(tarih) FROM etiket_snapshot WHERE tarih < CURRENT_DATE")
                date_row = cursor.fetchone()
                trend_tarihi = str(date_row[0]) if date_row and date_row[0] else None

                for row in prev_rows:
                    kolon = db_engine.val(row, 'etiket_kolon', '')
                    prev_sayi = db_engine.val(row, 'sayi', 0)
                    curr_sayi = (counts.get(kolon) or 0) if hasattr(counts, 'get') else 0
                    if prev_sayi > 0:
                        pct = round((curr_sayi - prev_sayi) / prev_sayi * 100, 1)
                    else:
                        pct = 100.0 if curr_sayi > 0 else 0
                    trend_map[kolon] = {
                        'degisim_yuzde': abs(pct),
                        'degisim_yonu': 'yukselis' if pct > 0 else ('dusus' if pct < 0 else 'sabit'),
                        'onceki_sayi': prev_sayi
                    }
        except Exception:
            pass  # etiket_snapshot tablosu yoksa veya boşsa trend gösterme

        # 4. Sonuçları Gruplandır
        etiketler_list = []
        kategoriler = {}

        for group_name, labels in LABEL_GROUPS.items():
            group_data = []
            for col in labels:
                count = counts.get(col) or 0
                item = {
                    'kolon': col,
                    'sayi': count,
                    'oran': round(count / total * 100, 1),
                    'trend': trend_map.get(col)
                }
                group_data.append(item)
                etiketler_list.append(item)
            kategoriler[group_name] = group_data

        result = {
            'toplam_musteri': total,
            'etiketler': etiketler_list,
            'kategoriler': kategoriler,
            'trend_tarihi': trend_tarihi
        }
        if not has_filter:
            _etiket_ozeti_cache[cache_key] = {'data': result, 'ts': time.time()}
        return Response(result)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"get_musteri_etiket_ozeti hatası: {e}")
        return Response({'error': str(e)}, status=500)
    finally:
        db_engine.release_connection(conn)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_musteri_zaman_cizelgesi(request, data_source_id, customer_id):
    """Müşteri zaman çizelgesi: aylık harcama + ziyaret + RFM segment geçmişi"""
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz'}, status=401)

    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()

        # 1. Aylık harcama ve ziyaret (satislar tablosundan)
        ay_expr = db_engine.strftime_expr('%Y-%m', 'tarih')
        cursor.execute(f"""
            SELECT {ay_expr} as ay,
                   COUNT(DISTINCT fis_no) as ziyaret_sayisi,
                   ROUND(SUM(tutar)::NUMERIC, 2) as toplam_tutar,
                   COUNT(DISTINCT urun_id) as urun_cesidi
            FROM satislar
            WHERE musteri_id = {ph}
            GROUP BY ay
            ORDER BY ay
        """, (customer_id,))
        aylik_ozet_raw = cursor.fetchall()

        # 2. RFM segment geçmişi (rfm_segment_log tablosundan — yoksa boş döner)
        # Tablo yoksa SAVEPOINT ile transaction'ı koruyoruz
        segment_gecmisi = []
        try:
            # Önce tablo var mı kontrol et
            cursor.execute("""
                SELECT EXISTS (
                    SELECT 1 FROM information_schema.tables
                    WHERE table_name = 'rfm_segment_log'
                ) as var
            """)
            tablo_var = cursor.fetchone()['var']
            if tablo_var:
                cursor.execute(f"""
                    SELECT {db_engine.strftime_expr('%Y-%m', 'kayit_tarihi')} as ay,
                           {db_engine.strftime_expr('%Y-%m-%d', 'kayit_tarihi')} as kayit_tarihi,
                           rfm_segment
                    FROM rfm_segment_log
                    WHERE musteri_id = {ph}
                    ORDER BY kayit_tarihi
                """, (customer_id,))
                segment_gecmisi = [dict(r) for r in cursor.fetchall()]
        except Exception:
            pass  # Hiçbir durumda transaction'ı bozmaz

        # 3. Segment map: ay → segment
        seg_map = {r['ay']: r['rfm_segment'] for r in segment_gecmisi}

        # 4. Aylık özeti serialize et, segment bilgisiyle zenginleştir
        aylik_ozet = []
        for r in aylik_ozet_raw:
            row = dict(r)
            row['rfm_segment'] = seg_map.get(row['ay'])
            # Decimal → float
            if hasattr(row.get('toplam_tutar'), 'real'):
                row['toplam_tutar'] = float(row['toplam_tutar'])
            aylik_ozet.append(row)

        # 5. Müşteri özet bilgileri — frontend'in beklediği key yapısına uygun
        cursor.execute(f"""
            SELECT ad_soyad, rfm_segment, ilk_alisveris_tarihi, son_alisveris_tarihi,
                   toplam_harcama, toplam_alisveris
            FROM musteridetayozet WHERE musteri_id = {ph}
        """, (customer_id,))
        ozet_row = cursor.fetchone()

        toplam_tutar = 0.0
        ilk_ay = None
        aktif_ay_sayisi = len(aylik_ozet)
        toplam_ziyaret = sum(r.get('ziyaret_sayisi', 0) or 0 for r in aylik_ozet)

        if ozet_row:
            raw_ozet = dict(ozet_row)
            toplam_tutar = float(raw_ozet.get('toplam_harcama') or 0)
            ilk_t = raw_ozet.get('ilk_alisveris_tarihi')
            if ilk_t:
                ilk_ay = str(ilk_t)[:7]  # 'YYYY-MM'

        ozet = {
            'toplam_tutar': toplam_tutar,
            'toplam_ziyaret': toplam_ziyaret,
            'ilk_ay': ilk_ay,
            'aktif_ay_sayisi': aktif_ay_sayisi,
        }

        return Response({
            'aylik_ozet': aylik_ozet,
            'segment_gecmisi': segment_gecmisi,
            'ozet': ozet,
        })

    except Exception as e:
        logger.error(f"Zaman çizelgesi hatası: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_segment_gecis_matrisi(request, data_source_id):
    """Segment geçiş matrisi: önceki ay vs bu ay segmentler arası müşteri akışı"""
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz'}, status=401)

    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)

        # rfm_segment_log tablosundan son iki distinct ay'ı bul
        try:
            cursor.execute("""
                SELECT DISTINCT {db_engine.strftime_expr('%Y-%m', 'kayit_tarihi')} as ay
                FROM rfm_segment_log
                ORDER BY ay DESC
                LIMIT 2
            """)
            aylar = [r['ay'] for r in cursor.fetchall()]
        except Exception:
            return Response({'matris': [], 'onceki_ay': None, 'bu_ay': None, 'mesaj': 'Henüz segment log verisi yok. İlk RFM güncellemesinden sonra görünür.'})

        if len(aylar) < 2:
            # Log var ama sadece 1 ay — mevcut segmentlerden statik görünüm döndür
            cursor.execute("""
                SELECT rfm_segment, COUNT(DISTINCT musteri_id) as musteri_sayisi
                FROM musteriler
                WHERE rfm_segment IS NOT NULL
                GROUP BY rfm_segment
            """)
            tek_ay = [dict(r) for r in cursor.fetchall()]
            return Response({
                'matris': [],
                'onceki_ay': None,
                'bu_ay': aylar[0] if aylar else None,
                'mevcut_dagilim': tek_ay,
                'mesaj': 'Geçiş matrisi için en az 2 aylık log gerekli.'
            })

        bu_ay, onceki_ay = aylar[0], aylar[1]

        # Geçiş matrisi: onceki_ay segmenti → bu_ay segmenti
        cursor.execute("""
            SELECT
                onceki.rfm_segment as kaynak_segment,
                buay.rfm_segment as hedef_segment,
                COUNT(DISTINCT onceki.musteri_id) as musteri_sayisi
            FROM rfm_segment_log onceki
            JOIN rfm_segment_log buay
                ON onceki.musteri_id = buay.musteri_id
            WHERE {db_engine.strftime_expr('%Y-%m', 'onceki.kayit_tarihi')} = {ph}
              AND {db_engine.strftime_expr('%Y-%m', 'buay.kayit_tarihi')} = {ph}
            GROUP BY kaynak_segment, hedef_segment
            ORDER BY musteri_sayisi DESC
        """, (onceki_ay, bu_ay))
        matris_rows = [dict(r) for r in cursor.fetchall()]

        # Segment listesi
        segments = sorted(set(
            [r['kaynak_segment'] for r in matris_rows] +
            [r['hedef_segment'] for r in matris_rows]
        ))

        return Response({
            'matris': matris_rows,
            'segments': segments,
            'onceki_ay': onceki_ay,
            'bu_ay': bu_ay,
        })

    except Exception as e:
        logger.error(f"Segment geçiş matrisi hatası: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_kategori_terk_listesi(request, data_source_id):
    """
    Kategori terk eden müşteriler — cache_kategori_terk tablosundan okunur.
    Cache yoksa anlık hesaplar (fallback).
    """
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz'}, status=401)

    ph = db_engine.ph()
    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)

        # Cache'den oku
        try:
            cursor.execute(
                "SELECT veri, hesaplama_tarihi FROM cache_kategori_terk ORDER BY hesaplama_tarihi DESC LIMIT 1"
            )
            row = cursor.fetchone()
            if row:
                import json
                veri = db_engine.val(row, 'veri')
                tarih = db_engine.val(row, 'hesaplama_tarihi')
                if isinstance(veri, str):
                    veri = json.loads(veri)
                if veri:
                    # Frontend'in beklediği key: 'musteriler' (cache'de 'terk_listesi')
                    return Response({
                        'musteriler': veri.get('terk_listesi', []),
                        'toplam': veri.get('toplam', 0),
                        '_cache_tarihi': str(tarih)[:19] if tarih else None,
                    })
        except Exception:
            pass

        # ── Fallback ────────────────────────────────────────────────────────
        logger.warning("cache_kategori_terk boş — anlık hesaplama yapılıyor.")
        limit = int(request.GET.get('limit', 50))

        cursor.execute(f"""
            SELECT
                m.id, m.ad, m.rfm_segment,
                COALESCE(md.toplam_harcama, 0) as toplam_harcama,
                COALESCE(dk.terk_edilen_kategori, 0) as terk_edilen_kategori_sayisi,
                COALESCE(dk.harcama_degisim_3ay_yuzde, 0) as harcama_degisim_3ay,
                COALESCE(dk.ziyaret_degisim_3ay_yuzde, 0) as ziyaret_degisim_3ay
            FROM musteriler m
            JOIN musterietiketler me ON m.id = me.musteri_id
            LEFT JOIN musteridetayozet md ON m.id = md.musteri_id
            LEFT JOIN musteridonem_karsilastirma dk ON m.id = dk.musteri_id
            WHERE me.kategori_terk_eden = TRUE
            ORDER BY terk_edilen_kategori_sayisi DESC, toplam_harcama DESC
            LIMIT {ph}
        """, (limit,))
        musteriler = [dict(r) for r in cursor.fetchall()]

        cursor.execute("""
            SELECT COUNT(DISTINCT musteri_id) as toplam
            FROM musterietiketler me
            JOIN musteriler m ON m.id = me.musteri_id
            WHERE me.kategori_terk_eden = TRUE
        """)
        toplam_row = cursor.fetchone()
        toplam = db_engine.val(toplam_row, 'toplam', 0) if toplam_row else 0

        return Response({'musteriler': musteriler, 'toplam': toplam, '_cache_tarihi': None})

    except Exception as e:
        logger.error(f"Kategori terk listesi hatası: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_kategori_terk_by_kategori(request, data_source_id):
    """
    Kategori terk eden müşterileri ana kategori bazında gruplar.
    - Parametresiz: {kategoriler: [{ana_kategori, musteri_sayisi}]}
    - ?ana_kategori=X&page=1&limit=20: paginated müşteri listesi
    - ?ana_kategori=X&format=xlsx: Excel export
    """
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz'}, status=401)
    if not validate_data_source(user, data_source_id):
        return Response({'error': 'Veri kaynağı bulunamadı veya erişim izni yok'}, status=404)

    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()

        ana_kategori = request.GET.get('ana_kategori', '').strip()
        fmt = request.GET.get('format', 'json')

        # Müşteri filtreleri
        customer_type_kt = request.GET.get('customer_type') or request.GET.get('customerType')
        approval_status_kt = request.GET.get('approval_status') or request.GET.get('approvalStatus')
        region_kt = request.GET.get('region')

        musteri_extra_kt = ""
        musteri_params_kt = []
        if customer_type_kt:
            musteri_extra_kt += f" AND m.tip = {ph}"
            musteri_params_kt.append(customer_type_kt)
        if approval_status_kt:
            musteri_extra_kt += f" AND m.onay_durumu = {ph}"
            musteri_params_kt.append(approval_status_kt)
        if region_kt:
            musteri_extra_kt += f" AND m.kayit_magazasi IN (SELECT id::text FROM magazalar WHERE {db_engine.bolge_expr()} = {ph})"
            musteri_params_kt.append(region_kt)

        if not ana_kategori:
            # Özet: her ana kategori için kaç terk eden müşteri var
            cursor.execute(f"""
                SELECT mkd.ana_kategori, COUNT(DISTINCT m.id) as musteri_sayisi
                FROM musteriler m
                JOIN musterietiketler me ON m.id = me.musteri_id
                JOIN musterikategoridagilimi mkd ON m.id = mkd.musteri_id
                WHERE me.kategori_terk_eden = TRUE
                  {musteri_extra_kt}
                GROUP BY mkd.ana_kategori
                ORDER BY musteri_sayisi DESC
            """, musteri_params_kt)
            kategoriler = [dict(r) for r in cursor.fetchall()]
            return Response({'kategoriler': kategoriler})

        # Detay: belirli bir kategorideki müşteriler
        try:
            page = max(1, int(request.GET.get('page', 1)))
        except (ValueError, TypeError):
            page = 1
        try:
            limit = min(100, max(1, int(request.GET.get('limit', 20))))
        except (ValueError, TypeError):
            limit = 20
        offset = (page - 1) * limit

        cursor.execute(f"""
            SELECT COUNT(DISTINCT m.id) as toplam
            FROM musteriler m
            JOIN musterietiketler me ON m.id = me.musteri_id
            JOIN musterikategoridagilimi mkd ON m.id = mkd.musteri_id
            WHERE me.kategori_terk_eden = TRUE AND mkd.ana_kategori = {ph}
              {musteri_extra_kt}
        """, [ana_kategori] + musteri_params_kt)
        toplam_row = cursor.fetchone()
        toplam = db_engine.val(toplam_row, 'toplam', 0)

        cursor.execute(f"""
            SELECT m.id, m.ad, m.rfm_segment,
                   COALESCE(md.toplam_harcama, 0) as toplam_harcama,
                   COALESCE(dk.terk_edilen_kategori, 0) as terk_edilen_kategori_sayisi,
                   COALESCE(dk.harcama_degisim_3ay_yuzde, 0) as harcama_degisim_3ay
            FROM musteriler m
            JOIN musterietiketler me ON m.id = me.musteri_id
            JOIN musterikategoridagilimi mkd ON m.id = mkd.musteri_id
            LEFT JOIN musteridetayozet md ON m.id = md.musteri_id
            LEFT JOIN musteridonem_karsilastirma dk ON m.id = dk.musteri_id
            WHERE me.kategori_terk_eden = TRUE AND mkd.ana_kategori = {ph}
              {musteri_extra_kt}
            GROUP BY m.id, m.ad, m.rfm_segment, md.toplam_harcama,
                     dk.terk_edilen_kategori, dk.harcama_degisim_3ay_yuzde
            ORDER BY toplam_harcama DESC
            LIMIT {ph} OFFSET {ph}
        """, [ana_kategori] + musteri_params_kt + [limit, offset])
        musteriler = [dict(r) for r in cursor.fetchall()]

        if fmt == 'xlsx':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from django.http import HttpResponse

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Kategori Terk'

                headers = ['Müşteri Adı', 'RFM Segment', 'Toplam Harcama (₺)', 'Terk Edilen Kat. Sayısı', 'Harcama Değ. 3 Ay (%)']
                header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                # Excel için tüm veriyi çek (limit yok)
                cursor.execute(f"""
                    SELECT m.id, m.ad, m.rfm_segment,
                           COALESCE(md.toplam_harcama, 0) as toplam_harcama,
                           COALESCE(dk.terk_edilen_kategori, 0) as terk_edilen_kategori_sayisi,
                           COALESCE(dk.harcama_degisim_3ay_yuzde, 0) as harcama_degisim_3ay
                    FROM musteriler m
                    JOIN musterietiketler me ON m.id = me.musteri_id
                    JOIN musterikategoridagilimi mkd ON m.id = mkd.musteri_id
                    LEFT JOIN musteridetayozet md ON m.id = md.musteri_id
                    LEFT JOIN musteridonem_karsilastirma dk ON m.id = dk.musteri_id
                    WHERE me.kategori_terk_eden = TRUE AND mkd.ana_kategori = {ph}
                    GROUP BY m.id, m.ad, m.rfm_segment, md.toplam_harcama,
                             dk.terk_edilen_kategori, dk.harcama_degisim_3ay_yuzde
                    ORDER BY toplam_harcama DESC
                """, (ana_kategori,))
                tum_musteriler = cursor.fetchall()

                for row_idx, m in enumerate(tum_musteriler, 2):
                    ws.cell(row=row_idx, column=1, value=db_engine.val(m, 'ad', ''))
                    ws.cell(row=row_idx, column=2, value=db_engine.val(m, 'rfm_segment', ''))
                    ws.cell(row=row_idx, column=3, value=round(float(db_engine.val(m, 'toplam_harcama', 0)), 2))
                    ws.cell(row=row_idx, column=4, value=int(db_engine.val(m, 'terk_edilen_kategori_sayisi', 0)))
                    ws.cell(row=row_idx, column=5, value=round(float(db_engine.val(m, 'harcama_degisim_3ay', 0)), 2))

                for col in ws.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                import io
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                safe_kat = ana_kategori.replace(' ', '_').replace('/', '-')
                response = HttpResponse(
                    buf.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="kategori_terk_{safe_kat}.xlsx"'
                return response

            except ImportError:
                return Response({'error': 'openpyxl kütüphanesi yüklü değil'}, status=500)

        return Response({'musteriler': musteriler, 'toplam': int(toplam), 'page': page, 'limit': limit})

    except Exception as e:
        logger.error(f"Kategori terk by kategori hatası: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)
