"""
Kohort, Ürün Birliktelik ve Marka Sadakati View'ları
----------------------------------------------------
Kohort / Enflasyon / Rakip / Hane analizleri artık cache tablolarından okunur.
Cache her gün analytics_cache.py tarafından doldurulur.
Cache boşsa (ilk çalıştırma veya hata), anlık hesaplama yapılır.
"""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
import logging
from .base import get_user_from_request, _read_cache
from .. import db_engine

logger = logging.getLogger(__name__)


# Imported from base.py: _read_cache(cursor, tablo) -> dict | None


# ─── Kohort Analizi ──────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_kohort_analizi(request, data_source_id):
    """
    Kohort retention matrisi — cache_kohort_analizi tablosundan okunur.
    Cache yoksa anlık hesaplama yapar (fallback).
    """
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz'}, status=401)

    customer_type = request.GET.get('customer_type') or request.GET.get('customerType')
    approval_status = request.GET.get('approval_status') or request.GET.get('approvalStatus')
    region = request.GET.get('region')
    has_filter = bool(customer_type or approval_status or region)

    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)

        # Cache'den oku (filtre yoksa)
        if not has_filter:
            cached = _read_cache(cursor, 'cache_kohort_analizi')
            if cached:
                return Response(cached)

        # ── Fallback: anlık hesapla ──────────────────────────────────────────
        if not has_filter:
            logger.warning("cache_kohort_analizi boş — anlık hesaplama yapılıyor.")
        from collections import defaultdict
        max_ay = int(request.GET.get('max_ay', 18))
        ph = db_engine.ph()

        # Müşteri filtresi oluştur
        musteri_filter = ""
        musteri_params = []
        if has_filter:
            filter_parts = []
            if customer_type:
                filter_parts.append(f"m.tip = {ph}")
                musteri_params.append(customer_type)
            if approval_status:
                filter_parts.append(f"m.onay_durumu = {ph}")
                musteri_params.append(approval_status)
            if region:
                filter_parts.append(f"m.kayit_magazasi IN (SELECT id::text FROM magazalar WHERE bolge = {ph})")
                musteri_params.append(region)
            musteri_filter = f"JOIN musteriler m ON md.musteri_id = m.id WHERE md.ilk_alisveris_tarihi IS NOT NULL AND {' AND '.join(filter_parts)}"
        else:
            musteri_filter = "WHERE md.ilk_alisveris_tarihi IS NOT NULL"

        cursor.execute(f"""
            SELECT md.musteri_id,
                   {db_engine.strftime_expr('%Y-%m', 'md.ilk_alisveris_tarihi')} as kohort_ay
            FROM musteridetayozet md
            {musteri_filter}
        """, musteri_params)
        rows = cursor.fetchall()
        if not rows:
            return Response({'kohortlar': [], 'max_ay': 0, 'mesaj': 'Müşteri verisi bulunamadı'})

        musteri_kohort = {}
        kohort_boyutlari = {}
        for r in rows:
            mid = db_engine.val(r, 'musteri_id')
            k = db_engine.val(r, 'kohort_ay')
            if mid and k:
                musteri_kohort[mid] = k
                kohort_boyutlari[k] = kohort_boyutlari.get(k, 0) + 1

        cursor.execute("""
            SELECT musteri_id, {db_engine.strftime_expr('%Y-%m', 'tarih')} as alis_ay
            FROM satislar
            WHERE musteri_id IS NOT NULL AND tarih IS NOT NULL
            GROUP BY musteri_id, {db_engine.strftime_expr('%Y-%m', 'tarih')}
        """)
        kohort_aktivite = defaultdict(lambda: defaultdict(set))
        for r in cursor.fetchall():
            mid = db_engine.val(r, 'musteri_id')
            alis_ay = db_engine.val(r, 'alis_ay')
            if mid is None or alis_ay is None:
                continue
            kohort_ay = musteri_kohort.get(mid)
            if not kohort_ay:
                continue
            try:
                ky, km = int(kohort_ay[:4]), int(kohort_ay[5:7])
                ay, am = int(alis_ay[:4]), int(alis_ay[5:7])
                indeks = (ay - ky) * 12 + (am - km)
                if 0 <= indeks <= max_ay:
                    kohort_aktivite[kohort_ay][indeks].add(mid)
            except (ValueError, IndexError):
                continue

        sonuc = []
        for kohort_ay in sorted(kohort_aktivite.keys()):
            boyut = kohort_boyutlari.get(kohort_ay, 0)
            if boyut == 0:
                continue
            retention = {
                ay_idx: round(len(kohort_aktivite[kohort_ay].get(ay_idx, set())) / boyut * 100, 1)
                for ay_idx in range(max_ay + 1)
            }
            sonuc.append({'kohort_ay': kohort_ay, 'kohort_boyutu': boyut, 'retention': retention})

        return Response({'kohortlar': sonuc[-24:], 'max_ay': max_ay, 'toplam_kohort': len(sonuc), '_cache_tarihi': None})

    except Exception as e:
        logger.error(f"Kohort analizi hatası: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)


# ─── Ürün Birliktelik (zaten hızlı tablo okuması — cache gerekmez) ────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_urun_birliktelik(request, data_source_id):
    """
    Ürün birliktelik analizi: lift, confidence, ortak_fis_sayisi ile sıralı liste.
    urunbirliktelikleri tablosu zaten önceden hesaplanmış — doğrudan okunur.
    """
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz'}, status=401)

    ph = db_engine.ph()
    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)

        min_lift = float(request.GET.get('min_lift', 1.0))
        min_confidence = float(request.GET.get('min_confidence', 0.0))
        limit = int(request.GET.get('limit', 100))
        sort_by = request.GET.get('sort_by', 'lift')
        sort_col = 'b.lift' if sort_by == 'lift' else ('b.confidence' if sort_by == 'confidence' else 'b.ortak_fis_sayisi')

        cursor.execute(f"""
            SELECT
                u1.id as urun1_id, u1.ad as urun1_ad,
                COALESCE(m1.ad, '') as urun1_marka,
                u2.id as urun2_id, u2.ad as urun2_ad,
                COALESCE(m2.ad, '') as urun2_marka,
                b.ortak_fis_sayisi, b.confidence, b.lift
            FROM urunbirliktelikleri b
            JOIN urunler u1 ON b.urun_id_1 = u1.id
            JOIN urunler u2 ON b.urun_id_2 = u2.id
            LEFT JOIN markalar m1 ON u1.marka_id = m1.id
            LEFT JOIN markalar m2 ON u2.marka_id = m2.id
            WHERE b.lift >= {ph}
              AND b.confidence >= {ph}
            ORDER BY {sort_col} DESC
            LIMIT {ph}
        """, (min_lift, min_confidence, limit))
        birliktelikler = [dict(r) for r in cursor.fetchall()]

        cursor.execute(f"""
            SELECT COUNT(DISTINCT urun_id) as toplam FROM urunbirliktelikleri
            WHERE lift >= {ph} AND confidence >= {ph}
        """, (min_lift, min_confidence))
        toplam_row = cursor.fetchone()
        toplam = db_engine.val(toplam_row, 'toplam', 0) if toplam_row else 0

        cursor.execute("""
            SELECT
                k1.ana as kat1, k2.ana as kat2,
                 COUNT(DISTINCT kural) as kural_sayisi,
                AVG(b.lift) as ort_lift,
                SUM(b.ortak_fis_sayisi) as toplam_fis
            FROM grupbirliktelikleri b
            JOIN kategoriler k1 ON b.kategori_id_1 = k1.id
            JOIN kategoriler k2 ON b.kategori_id_2 = k2.id
            WHERE b.tip = 'CAT_ONLY_SQL'
            GROUP BY k1.ana, k2.ana
            ORDER BY toplam_fis DESC
            LIMIT 20
        """)
        kat_birliktelik = [dict(r) for r in cursor.fetchall()]

        return Response({
            'birliktelikler': birliktelikler,
            'toplam': toplam,
            'kategori_birliktelik': kat_birliktelik,
        })

    except Exception as e:
        logger.error(f"Ürün birliktelik hatası: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)


# ─── Marka Sadakati ──────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_marka_sadakati(request, data_source_id):
    """
    Marka sadakati analizi — cache_marka_sadakati tablosundan okunur.
    Cache yoksa anlık hesaplar (fallback).
    """
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz'}, status=401)

    customer_type_ms = request.GET.get('customer_type') or request.GET.get('customerType')
    approval_status_ms = request.GET.get('approval_status') or request.GET.get('approvalStatus')
    region_ms = request.GET.get('region')
    has_filter_ms = bool(customer_type_ms or approval_status_ms or region_ms)

    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)

        if not has_filter_ms:
            cached = _read_cache(cursor, 'cache_marka_sadakati')
            if cached:
                return Response(cached)

        # ── Fallback: marka verileri hesapla ──
        if not has_filter_ms:
            logger.warning("cache_marka_sadakati boş — anlık hesaplama yapılıyor.")

        ph = db_engine.ph()
        # Müşteri filtresi
        musteri_join_ms = ""
        musteri_params_ms = []
        if has_filter_ms:
            filter_parts_ms = []
            if customer_type_ms:
                filter_parts_ms.append(f"mu.tip = {ph}")
                musteri_params_ms.append(customer_type_ms)
            if approval_status_ms:
                filter_parts_ms.append(f"mu.onay_durumu = {ph}")
                musteri_params_ms.append(approval_status_ms)
            if region_ms:
                filter_parts_ms.append(f"mu.kayit_magazasi IN (SELECT id::text FROM magazalar WHERE bolge = {ph})")
                musteri_params_ms.append(region_ms)
            musteri_join_ms = f"AND {' AND '.join(filter_parts_ms)}"

        musteri_where_ms = f"WHERE md.marka_adi IS NOT NULL {musteri_join_ms}"

        cursor.execute(f"""
            SELECT
                md.marka_adi as marka,
                COUNT(DISTINCT md.musteri_id) as musteri_sayisi,
                SUM(md.toplam_harcama) as toplam_harcama,
                AVG(md.toplam_harcama) as ort_harcama,
                COUNT(DISTINCT CASE WHEN mu.rfm_segment LIKE '%Şampiyon%' THEN md.musteri_id END) as sampiyonlar,
                COUNT(DISTINCT CASE WHEN mu.rfm_segment LIKE '%Sadık%' THEN md.musteri_id END) as sadik,
                COUNT(DISTINCT CASE WHEN mu.rfm_segment LIKE '%Risk%' OR mu.rfm_segment LIKE '%Kayıp%' THEN md.musteri_id END) as risk
            FROM musterimarka_dagilimi md
            JOIN musteriler mu ON md.musteri_id = mu.id
            {musteri_where_ms}
            GROUP BY md.marka_adi
            ORDER BY toplam_harcama DESC
            LIMIT 30
        """, musteri_params_ms)
        marka_rows = [dict(r) for r in cursor.fetchall()]

        cursor.execute(f"""
            SELECT
                md.marka_adi as marka,
                COUNT(DISTINCT md.musteri_id) as toplam_musteri,
                COUNT(DISTINCT CASE WHEN alt.marka_sayisi = 1 THEN md.musteri_id END) as sadece_bu_marka
            FROM musterimarka_dagilimi md
            JOIN musteriler mu ON md.musteri_id = mu.id
            JOIN (
                SELECT musteri_id, COUNT(DISTINCT marka_adi) as marka_sayisi
                FROM musterimarka_dagilimi GROUP BY musteri_id
            ) alt ON md.musteri_id = alt.musteri_id
            {musteri_where_ms}
            GROUP BY md.marka_adi
            ORDER BY toplam_musteri DESC
            LIMIT 30
        """, musteri_params_ms)
        sadakat_rows = [dict(r) for r in cursor.fetchall()]

        cursor.execute(f"""
            SELECT md.marka_adi as marka, COUNT(DISTINCT md.musteri_id) as musteri_sayisi,
                   ROUND(AVG(md.toplam_harcama), 0) as ort_musteri_harcama
            FROM musterimarka_dagilimi md
            JOIN musteriler mu ON md.musteri_id = mu.id
            {musteri_where_ms}
            GROUP BY md.marka_adi ORDER BY musteri_sayisi DESC LIMIT 30
        """, musteri_params_ms)
        top_markalar = [dict(r) for r in cursor.fetchall()]

        return Response({
            'marka_profiller': marka_rows,
            'sadakat_skorlari': sadakat_rows,
            'top_markalar': top_markalar,
            '_cache_tarihi': None,
        })

    except Exception as e:
        logger.error(f"Marka sadakati hatası: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_cohort_analysis_excel(request, data_source_id):
    """Kohort analizini Excel olarak dışa aktar"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    # get_kohort_analizi mantığını kullan ama Response yerine veri al
    res = get_kohort_analizi(request._request, data_source_id)
    if res.status_code != 200:
        return HttpResponse("Veri çekilemedi", status=res.status_code)
    
    data = res.data
    kohortlar = data.get('kohortlar', [])
    max_ay = data.get('max_ay', 12)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kohort Analizi"
    
    # Başlıklar
    headers = ["Kohort Ayı", "Müşteri Sayısı"] + [f"Ay {i}" for i in range(max_ay + 1)]
    ws.append(headers)
    
    # Stil
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for k in kohortlar:
        row = [k['kohort_ay'], k['kohort_boyutu']]
        ret = k.get('retention', {})
        for i in range(max_ay + 1):
            row.append(f"%{ret.get(i, 0)}" if i in ret else "-")
        ws.append(row)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Kohort_Analizi_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_brand_loyalty_excel(request, data_source_id):
    """Marka sadakati analizini Excel olarak dışa aktar"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from django.http import HttpResponse
    
    res = get_marka_sadakati(request._request, data_source_id)
    if res.status_code != 200:
        return HttpResponse("Veri çekilemedi", status=res.status_code)
        
    data = res.data
    profiller = data.get('marka_profiller', [])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marka Sadakati"
    
    headers = ["Marka", "Müşteri Sayısı", "Toplam Harcama", "Ort. Harcama", "Şampiyonlar", "Sadık", "Risk Grubu"]
    ws.append(headers)
    
    for p in profiller:
        ws.append([
            p['marka'], p['musteri_sayisi'], p['toplam_harcama'], 
            p['ort_harcama'], p['sampiyonlar'], p['sadik'], p['risk']
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Marka_Sadakati_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
