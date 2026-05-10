from django.http import HttpResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
import io
import json
import re as _re
import logging

from .. import db_engine

logger = logging.getLogger(__name__)

# ── Module-level Excel style helpers (canonical definitions) ─────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_XL_TEXT_DARK = '1E293B'
_XL_GREY_MID = 'E2E8F0'

def _xl_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _xl_font(bold=False, color=_XL_TEXT_DARK, size=11, italic=False):
    return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)

def _xl_border_bottom(color=_XL_GREY_MID):
    s = Side(style='thin', color=color)
    return Border(bottom=s)

def _xl_border_all(color=_XL_GREY_MID):
    s = Side(style='thin', color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def _xl_align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
# ──────────────────────────────────────────────────────────────────────────


def _parse_onerilen_urunler(camp):
    urunler = []
    raw = db_engine.val(camp, 'onerilen_urunler')
    if raw:
        try:
            parsed = json.loads(raw) if isinstance(raw, str) else raw
            if isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict):
                        urunler.append(item.get('ad', ''))
                    elif isinstance(item, str):
                        urunler.append(item)
        except (json.JSONDecodeError, TypeError):
            pass
    if not urunler:
        fallback = db_engine.val(camp, 'urun_ad') or db_engine.val(camp, 'kategori_ad') or ''
        if fallback:
            urunler.append(fallback)
    return urunler


def _build_excel_response(oneri_ids):
    from datetime import datetime
    from openpyxl import Workbook

    COLUMNS = [
        'Kampanya Tipi', 'Müşterinin Aldığı Kategori', 'Önerilen Kategori',
        'Onay Durumu', 'Müşteri ID', 'Müşteri Ad Soyad', 'Telefon',
        'RFM Segment', 'Müşteri Tipi',
        'Önerilen Ürün 1', 'Önerilen Ürün 2', 'Önerilen Ürün 3', 'Önerilen Ürün 4', 'Önerilen Ürün 5'
    ]

    import time as _time
    conn = db_engine.get_connection()
    cursor = db_engine.get_dict_cursor(conn)
    ph = db_engine.ph()

    try:
        _t0 = _time.time()
        logger.info(f"[EXPORT] istenen_ids={oneri_ids}")
        # ── 1. Tüm kampanyaları tek sorguda çek ─────────────────────────────
        placeholders = ','.join([ph] * len(oneri_ids))
        cursor.execute(
            f"SELECT * FROM otomatikkampanyaonerileri WHERE oneri_id IN ({placeholders})",
            oneri_ids
        )
        camps = cursor.fetchall()
        logger.info(f"[EXPORT] adım1 kampanyalar={len(camps)} t={_time.time()-_t0:.2f}s")

        # ── 2. Kampanyaları tipe göre ayır ──────────────────────────────────
        cross_sell_camps = []   # (camp, source_cat_id, target_cat_id)
        segment_camps = []      # (camp, segment_clean)

        for camp in camps:
            camp_tipi = db_engine.val(camp, 'kampanya_tipi', '')
            source_cat_id = db_engine.val(camp, 'ikinci_urun_id')
            target_cat_id = db_engine.val(camp, 'kategori_id')
            hedef_segment = db_engine.val(camp, 'hedef_segment', '')
            if camp_tipi == 'Cross-Sell' and source_cat_id:
                cross_sell_camps.append((camp, source_cat_id, target_cat_id))
            else:
                segment_clean = hedef_segment.replace(' Alıcıları', '').strip()
                segment_camps.append((camp, segment_clean))

        # ── 3. satislar'a hiç dokunmadan müşteri listesi ──────────────────────
        # Her kampanya için hedef_musteri_sayisi kadar müşteri musteriler'den çek
        # Segment kampanyalar: rfm_segment eşleşmesine bak (musteriler tablosu hızlı)
        # Cross-sell kampanyalar: hedef_segment = "X Alıcıları" — musteriler.rfm_segment ile eşleşmez
        #   → Bu kampanyalar için onaylı (veya tüm) müşterilerden limit'li liste al

        # Tüm kampanyalara ait hedef_segment'leri topla
        all_segments = []
        for camp in camps:
            seg = db_engine.val(camp, 'hedef_segment', '') or ''
            seg_clean = seg.replace(' Alıcıları', '').strip()
            all_segments.append(seg_clean)
        unique_segments_all = list({s for s in all_segments if s})

        # musteriler tablosundan fallback müşteri listesi çek (cross-sell için)
        try:
            _tm = _time.time()
            cursor.execute(
                "SELECT id, ad, telefon, rfm_segment, tip, onay_durumu FROM musteriler LIMIT 1000"
            )
            all_customers = cursor.fetchall()
            logger.info(f"[EXPORT] adım3 musteriler={len(all_customers)} t={_time.time()-_tm:.2f}s")
        except Exception as e:
            logger.error(f"Müşteri listesi sorgu hatası: {e}")
            all_customers = []

        # Segment → keyword eşleme tablosu
        # hedef_segment (ASCII-normalized, ör. "Sadiklar") → rfm_segment'te aranacak Türkçe anahtar kelime
        _SEG_KEYWORD_MAP = {
            'Sampiyonlar':             'Şampiyonlar',
            'Sadiklar':                'Sadık Müşteriler',
            'Sadik Musteriler':        'Sadık Müşteriler',
            'Yuksek Harcama Yapanlar': 'Yüksek Harcama',
            'Potansiyel Sampiyonlar':  'Potansiyel Şampiyonlar',
            'Tekrar Kazanilanlar':     'Tekrar Kazanılan',
            'Yeni Musteriler':         'Yeni Müşteriler',
            'Sadik Olmaya Adaylar':    'Sadık Olmaya Adaylar',
            'Ilgi Bekleyenler':        'İlgi Bekleyenler',
            'Kayip Musteriler':        'Kayıp Müşteriler',
            'Risk Altindakiler':       'Risk Altındakiler',
            'Uyuyanlar':               'Uyuyanlar',
        }

        # Segment → müşteri eşlemesi: her segment için musteriler'den doğrudan sorgula
        segment_customer_map = {}
        for seg in unique_segments_all:
            keyword = _SEG_KEYWORD_MAP.get(seg)
            if keyword:
                try:
                    cursor.execute(
                        f"SELECT id, ad, telefon, rfm_segment, tip, onay_durumu FROM musteriler WHERE rfm_segment LIKE {ph} LIMIT 500",
                        [f'%{keyword}%']
                    )
                    segment_customer_map[seg] = cursor.fetchall()
                    logger.info(f"[EXPORT] segment '{seg}' → keyword '{keyword}' → {len(segment_customer_map[seg])} müşteri")
                except Exception as e:
                    logger.error(f"Segment müşteri sorgu hatası ({seg}): {e}")
                    segment_customer_map[seg] = []
            else:
                # keyword bulunamadı, substring fallback dene
                segment_customer_map[seg] = [
                    c for c in all_customers
                    if seg.lower() in (db_engine.val(c, 'rfm_segment', '') or '').lower()
                ]
                logger.warning(f"[EXPORT] segment '{seg}' için keyword eşlemesi yok, fallback kullanıldı: {len(segment_customer_map[seg])} müşteri")

        # Cross-sell için: segment eşleşmeyenlere tüm listeyi ver (hedef_musteri_sayisi ile cap'li)
        cat_pairs = list({(sc, tc) for _, sc, tc in cross_sell_camps})
        customer_map = {}
        for source_cat_id, target_cat_id in cat_pairs:
            # Cross-sell için kaynak kategori adına göre eşleşen müşteri yoksa tüm listeyi al
            camp_match = next((c for c, sc, tc in cross_sell_camps if sc == source_cat_id and tc == target_cat_id), None)
            hedef_seg = db_engine.val(camp_match, 'hedef_segment', '') if camp_match else ''
            seg_clean = hedef_seg.replace(' Alıcıları', '').strip()
            matched = segment_customer_map.get(seg_clean, [])
            # rfm_segment eşleşmediyse (cross-sell "X Alıcıları" formatı), tüm listeden al
            if not matched:
                limit = int(db_engine.val(camp_match, 'hedef_musteri_sayisi', 500) or 500)
                matched = all_customers[:min(limit, 500)]
            customer_map[(source_cat_id, target_cat_id)] = matched

        logger.info(f"[EXPORT] adım4 customer_map hazır t={_time.time()-_t0:.2f}s")
        # ── 7. Excel oluştur ─────────────────────────────────────────────────
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter

        # Renk paleti
        C_PURPLE_DARK  = '3730A3'   # başlık arka planı (koyu mor)
        C_PURPLE_MID   = '6366F1'   # kampanya başlığı
        C_PURPLE_LIGHT = 'EEF2FF'   # mor açık arka plan
        C_GREEN_DARK   = '065F46'   # yeşil metin
        C_GREEN_LIGHT  = 'D1FAE5'   # yeşil açık arka plan
        C_ORANGE_LIGHT = 'FFF7ED'   # turuncu açık
        C_ORANGE_DARK  = 'C2410C'   # turuncu metin
        C_BLUE_LIGHT   = 'EFF6FF'   # mavi açık
        C_BLUE_MID     = '3B82F6'   # mavi orta
        C_GREY_LIGHT   = 'F8FAFC'   # gri açık
        C_GREY_MID     = 'E2E8F0'   # gri orta
        C_WHITE        = 'FFFFFF'
        C_TEXT_DARK    = '1E293B'
        C_TEXT_MID     = '475569'

        # Module-level helpers (defined at top of file)
        fill = _xl_fill
        font = _xl_font
        border_bottom = _xl_border_bottom
        border_all = _xl_border_all
        align = _xl_align

        wb = Workbook()
        wb.remove(wb.active)

        # ════════════════════════════════════════════════════════════════
        # SAYFA 1: Kampanya Özeti
        # ════════════════════════════════════════════════════════════════
        ws1 = wb.create_sheet('Kampanya Ozeti')
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions['A'].width = 52
        ws1.column_dimensions['B'].width = 45

        # Sayfa başlığı — büyük mor banner
        ws1.merge_cells('A1:B1')
        title_cell = ws1['A1']
        title_cell.value = 'KAMPANYA ANALIZ RAPORU'
        title_cell.font = Font(name='Calibri', bold=True, color=C_WHITE, size=16)
        title_cell.fill = fill(C_PURPLE_DARK)
        title_cell.alignment = align('center')
        ws1.row_dimensions[1].height = 36

        ws1.merge_cells('A2:B2')
        sub_cell = ws1['A2']
        sub_cell.value = f'Olusturulma Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        sub_cell.font = font(color='A5B4FC', size=10, italic=True)
        sub_cell.fill = fill(C_PURPLE_DARK)
        sub_cell.alignment = align('center')
        ws1.row_dimensions[2].height = 20

        current_row = 3

        for camp in camps:
            camp_tipi     = db_engine.val(camp, 'kampanya_tipi', '')
            kategori      = db_engine.val(camp, 'kategori_ad', '') or ''
            kaynak_kat    = db_engine.val(camp, 'kaynak_kategori_ad') or db_engine.val(camp, 'ikinci_urun_ad') or ''
            hedef_seg     = db_engine.val(camp, 'hedef_segment', '') or ''
            hedef_say     = int(db_engine.val(camp, 'hedef_musteri_sayisi', 0) or 0)
            fis_say       = int(db_engine.val(camp, 'fis_sayisi', 0) or 0)
            pot_ciro      = float(db_engine.val(camp, 'potansiyel_ciro', 0) or 0)
            tahmini_kat   = int(db_engine.val(camp, 'tahmini_katilim', 0) or 0) or max(1, round(hedef_say * 0.15))
            lift          = db_engine.val(camp, 'lift')
            guven         = db_engine.val(camp, 'guven')
            veri_ozeti        = db_engine.val(camp, 'veri_ozeti', '') or ''
            gerekcesi         = db_engine.val(camp, 'gerekcesi', '') or ''
            beklenen_sonuc    = db_engine.val(camp, 'beklenen_sonuc', '') or ''
            onerilen_indirim  = float(db_engine.val(camp, 'onerilen_indirim', 0) or 0)
            onerilen_min_tutar= float(db_engine.val(camp, 'onerilen_min_tutar', 0) or 0)
            gecerlilik_suresi = int(db_engine.val(camp, 'gecerlilik_suresi', 7) or 7)
            roi_tahmini_db    = float(db_engine.val(camp, 'roi_tahmini', 0) or 0)
            source_cat_id = db_engine.val(camp, 'ikinci_urun_id')
            target_cat_id = db_engine.val(camp, 'kategori_id')

            onerilen_raw = db_engine.val(camp, 'onerilen_urunler', '[]') or '[]'
            try:
                onerilen_urunler = json.loads(onerilen_raw) if isinstance(onerilen_raw, str) else onerilen_raw
            except Exception:
                onerilen_urunler = []

            ortak_musteri = 0
            if veri_ozeti:
                m2 = _re.search(r'Ortak m[üu][şs]teri:\s*(\d+)', veri_ozeti)
                if m2:
                    ortak_musteri = int(m2.group(1))

            urun_adlari = [u.get('ad', '') if isinstance(u, dict) else str(u) for u in (onerilen_urunler or [])]
            urun_listesi_str = ', '.join(filter(None, urun_adlari[:5])) or db_engine.val(camp, 'urun_ad', '') or ''

            # Boşluk satırı
            current_row += 1
            ws1.row_dimensions[current_row].height = 8

            # Kampanya başlık satırı
            current_row += 1
            ws1.merge_cells(f'A{current_row}:B{current_row}')
            if camp_tipi == 'Cross-Sell':
                title_text = f'CAPRAZ SATIS KAMPANYASI  |  {kaynak_kat}  ->  {kategori}'
            else:
                title_text = f'{camp_tipi.upper()} KAMPANYASI  |  {kategori or hedef_seg}'
            c = ws1.cell(row=current_row, column=1, value=title_text)
            c.font = Font(name='Calibri', bold=True, color=C_WHITE, size=12)
            c.fill = fill(C_PURPLE_MID)
            c.alignment = align('left')
            ws1.row_dimensions[current_row].height = 28

            # Kolon başlıkları
            current_row += 1
            for col, label in enumerate(['METRIK', 'DEGER'], start=1):
                c = ws1.cell(row=current_row, column=col, value=label)
                c.font = font(bold=True, color=C_WHITE, size=10)
                c.fill = fill('312E81')
                c.alignment = align('center' if col == 2 else 'left')
                c.border = border_all('312E81')
            ws1.row_dimensions[current_row].height = 22

            def add_row(label, value, row_fill=C_WHITE, label_bold=False, label_indent=0, value_is_number=False, value_color=C_TEXT_DARK, label_color=C_TEXT_MID, wrap_val=False, val_align='right'):
                nonlocal current_row
                current_row += 1
                indent = '  ' * label_indent
                ca = ws1.cell(row=current_row, column=1, value=indent + label)
                ca.font = font(bold=label_bold, color=label_color, size=10)
                ca.fill = fill(row_fill)
                ca.alignment = align('left', wrap=True)
                ca.border = border_bottom()

                cb = ws1.cell(row=current_row, column=2, value=value)
                cb.font = font(bold=label_bold, color=value_color, size=10)
                cb.fill = fill(row_fill)
                cb.alignment = align(val_align, wrap=wrap_val)
                cb.border = border_bottom()
                if value_is_number and isinstance(value, (int, float)):
                    cb.number_format = '#,##0.00'
                
                # Sadece wrap yapılmayacak kısa satırlarda sabit yükseklik ver, 
                # diğerlerinde excel otomatik ayarlasın (veya biraz daha büyük sabit değer)
                if wrap_val:
                    ws1.row_dimensions[current_row].height = 40
                else:
                    ws1.row_dimensions[current_row].height = 20

            if camp_tipi == 'Cross-Sell':
                lift_val  = round(float(lift), 2) if lift else 0
                guven_val = round(float(guven) * 100, 1) if guven else 0
                mevcut_birlikte_ciro = float(db_engine.val(camp, 'mevcut_birlikte_ciro', 0) or 0)
                birlikte_ciro        = float(db_engine.val(camp, 'birlikte_ciro', 0) or 0)
                # mevcut_birlikte_ciro 0 ise fis_sayisi × ürün ortalama fiyatlarından hesapla
                if mevcut_birlikte_ciro <= 0 and fis_say > 0 and onerilen_urunler:
                    avg_urun_fiyat = 0.0
                    prices = [float(u.get('ort', 0) or u.get('fiyat', 0) or 0) for u in onerilen_urunler if isinstance(u, dict)]
                    prices = [p for p in prices if p > 0]
                    if prices:
                        avg_urun_fiyat = sum(prices) / len(prices)
                    elif onerilen_min_tutar > 0:
                        avg_urun_fiyat = round(onerilen_min_tutar / 0.75, 2)
                    if avg_urun_fiyat > 0:
                        mevcut_birlikte_ciro = round(fis_say * avg_urun_fiyat, 2)
                kt = hedef_say

                kaynak_toplam = 0
                try:
                    cursor.execute("""
                        SELECT COUNT(DISTINCT s.musteri_id) as toplam
                        FROM satislar s
                        JOIN kategoriler k ON k.id = s.kategori_id
                        WHERE (k.alt1 = %s OR k.alt2 = %s OR k.ana = %s)
                          AND s.musteri_id IS NOT NULL AND s.miktar > 0 AND s.tutar > 0
                    """, [kaynak_kat, kaynak_kat, kaynak_kat])
                    r_src = cursor.fetchone()
                    kaynak_toplam = int(r_src['toplam']) if r_src else kt + ortak_musteri
                except:
                    kaynak_toplam = kt + ortak_musteri

                ind_oran = (onerilen_indirim or 25.0) / 100.0
                if ind_oran >= 1 or ind_oran <= 0: ind_oran = 0.25
                indirim_maliyeti = max(round(pot_ciro / (1 - ind_oran) * ind_oran, 2), 1)

                urun_roi = round(pot_ciro / indirim_maliyeti, 2)
                birlikte_roi = round(birlikte_ciro / indirim_maliyeti, 2)

                # Mor grup — hedef kitle
                add_row(f"'{kaynak_kat}' kategorisi toplam alicisi", f"{kaynak_toplam:,} musteri", C_PURPLE_LIGHT, label_color='4338CA')
                add_row(f"Henuz '{kategori}' almamis hedef kitle", f"{hedef_say:,} musteri", C_PURPLE_LIGHT, label_color='4338CA')
                add_row("Kampanyaya ulasilacak kitle (lift bazli)", f"{tahmini_kat:,} musteri", C_PURPLE_LIGHT, label_color='4338CA')
                # Mavi grup — fis verisi
                add_row("Bu birlikteligi gosteren islem (fis) sayisi", f"{fis_say:,} fis", C_BLUE_LIGHT, label_color='1D4ED8')
                add_row("Her iki kategoriyi birlikte alan musteri", f"{ortak_musteri:,} musteri", C_BLUE_LIGHT, label_color='1D4ED8')
                add_row("Lift skoru", f"{lift_val}x", C_BLUE_LIGHT, label_color='1D4ED8')
                add_row("Guven skoru", f"%{guven_val}", C_BLUE_LIGHT, label_color='1D4ED8')
                # Turuncu grup — mevcut ciro
                add_row(f"Mevcut birlikte alim cirosu ({fis_say:,} fis)", mevcut_birlikte_ciro, C_ORANGE_LIGHT, label_bold=True, value_is_number=True, value_color=C_ORANGE_DARK, label_color=C_ORANGE_DARK)
                # Yeşil grup — potansiyel
                add_row(f"'{kategori}' kampanyasi yatirim maliyeti (Indirim)", indirim_maliyeti, C_GREEN_LIGHT, label_bold=True, value_is_number=True, value_color=C_GREEN_DARK, label_color=C_GREEN_DARK)
                add_row(f"'{kategori}' onerilen urun potansiyel cirosu", pot_ciro, C_GREEN_LIGHT, label_bold=True, value_is_number=True, value_color=C_GREEN_DARK, label_color=C_GREEN_DARK)
                add_row(f"'{kategori}' onerilen urun tahmini ROI", f"{urun_roi}x", C_GREEN_LIGHT, label_bold=True, value_color=C_GREEN_DARK, label_color=C_GREEN_DARK)
                add_row("Potansiyel birlikte alim cirosu", birlikte_ciro if birlikte_ciro > 0 else pot_ciro, 'DCFCE7', label_bold=True, value_is_number=True, value_color='166534', label_color='166534')
                add_row("Birlikte tahmini ROI (Capraz Satis Etkisiyle)", f"{birlikte_roi}x", 'DCFCE7', label_bold=True, value_color='166534', label_color='166534')
                # Uygulama önerisi
                if onerilen_indirim > 0:
                    add_row("Onerilen indirim orani", f"%{round(onerilen_indirim)}", C_GREY_LIGHT, label_color=C_TEXT_MID)
                if onerilen_min_tutar > 0:
                    add_row("Onerilen kampanya fiyati", onerilen_min_tutar, C_GREY_LIGHT, value_is_number=True, label_color=C_TEXT_MID)
                add_row("Gecerlilik suresi", f"{gecerlilik_suresi} gun", C_GREY_LIGHT, label_color=C_TEXT_MID)
                # Ürün detayları
                for i, u in enumerate(onerilen_urunler[:5], 1):
                    ad  = u.get('ad', '') if isinstance(u, dict) else str(u)
                    # 'ort' veya 'fiyat' alanından fiyat al, fallback: onerilen_min_tutar / (1 - ind_oran)
                    if isinstance(u, dict):
                        fiy = float(u.get('ort', 0) or u.get('fiyat', 0) or 0)
                    else:
                        fiy = 0
                    if fiy <= 0 and onerilen_min_tutar > 0:
                        fiy = round(onerilen_min_tutar / (1 - ind_oran), 2)
                    if ad:
                        label = f"Urun {i}: {ad}"
                        if fiy > 0:
                            indirim_yuzdesi = round(onerilen_indirim) if onerilen_indirim > 0 else 25
                            kampanya_fiyati = round(fiy * (1 - ind_oran), 2)
                            val = f"₺{fiy:,.2f} -> ₺{kampanya_fiyati:,.2f} (%{indirim_yuzdesi} ind.)"
                        else:
                            val = '-'
                        add_row(label, val, C_GREY_LIGHT, label_color=C_TEXT_MID)
                # Gerekçe ve beklenen sonuç
                if gerekcesi:
                    add_row("Neden bu kampanya?", gerekcesi, 'FEF9C3', label_bold=True, label_color='92400E', wrap_val=True, val_align='left')
                if beklenen_sonuc:
                    add_row("Beklenen sonuc", beklenen_sonuc, 'FEF9C3', label_color='92400E', wrap_val=True, val_align='left')
                if veri_ozeti:
                    add_row("Veri kaynagi", veri_ozeti, C_GREY_LIGHT, label_color=C_TEXT_MID, wrap_val=True, val_align='left')
            else:
                add_row("Hedef kitle", f"{hedef_say:,} musteri", C_PURPLE_LIGHT, label_color='4338CA')
                add_row("Kampanyaya ulasilacak kitle", f"{tahmini_kat:,} musteri", C_PURPLE_LIGHT, label_color='4338CA')
                add_row("Potansiyel ciro", pot_ciro, C_GREEN_LIGHT, label_bold=True, value_is_number=True, value_color=C_GREEN_DARK, label_color=C_GREEN_DARK)
                if roi_tahmini_db > 0:
                    add_row("Tahmini ROI", f"{roi_tahmini_db}x", C_GREEN_LIGHT, label_bold=True, value_color=C_GREEN_DARK, label_color=C_GREEN_DARK)
                if onerilen_indirim > 0:
                    add_row("Onerilen indirim", f"%{round(onerilen_indirim)}", C_GREY_LIGHT, label_color=C_TEXT_MID)
                add_row("Gecerlilik suresi", f"{gecerlilik_suresi} gun", C_GREY_LIGHT, label_color=C_TEXT_MID)
                ind_oran_else = (onerilen_indirim or 25.0) / 100.0
                if ind_oran_else >= 1 or ind_oran_else <= 0: ind_oran_else = 0.25
                for i, u in enumerate(onerilen_urunler[:5], 1):
                    ad  = u.get('ad', '') if isinstance(u, dict) else str(u)
                    if isinstance(u, dict):
                        fiy = float(u.get('ort', 0) or u.get('fiyat', 0) or 0)
                    else:
                        fiy = 0
                    if fiy <= 0 and onerilen_min_tutar > 0:
                        fiy = round(onerilen_min_tutar / (1 - ind_oran_else), 2)
                    if ad:
                        if fiy > 0:
                            kampanya_fiyati_e = round(fiy * (1 - ind_oran_else), 2)
                            indirim_yuzdesi_e = round(onerilen_indirim) if onerilen_indirim > 0 else 25
                            val_e = f"₺{fiy:,.2f} -> ₺{kampanya_fiyati_e:,.2f} (%{indirim_yuzdesi_e} ind.)"
                        else:
                            val_e = '-'
                        add_row(f"Urun {i}: {ad}", val_e, C_GREY_LIGHT, label_color=C_TEXT_MID)
                if gerekcesi:
                    add_row("Neden bu kampanya?", gerekcesi, 'FEF9C3', label_bold=True, label_color='92400E', wrap_val=True, val_align='left')
                if beklenen_sonuc:
                    add_row("Beklenen sonuc", beklenen_sonuc, 'FEF9C3', label_color='92400E', wrap_val=True, val_align='left')
                if veri_ozeti:
                    add_row("Veri kaynagi", veri_ozeti, C_GREY_LIGHT, label_color=C_TEXT_MID, wrap_val=True, val_align='left')

        # ════════════════════════════════════════════════════════════════
        # SAYFA 2: Hedef Müşteri Listesi
        # ════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Hedef Kitle')
        ws2.sheet_view.showGridLines = False

        # Başlık satırı
        ws2.merge_cells('A1:N1')
        hdr = ws2['A1']
        hdr.value = 'HEDEF MUSTERI LISTESI'
        hdr.font = Font(name='Calibri', bold=True, color=C_WHITE, size=14)
        hdr.fill = fill(C_PURPLE_DARK)
        hdr.alignment = align('center')
        ws2.row_dimensions[1].height = 32

        # Kolon başlıkları
        col_headers = [
            'Kampanya Tipi', 'Kaynak Kategori', 'Onerilen Kategori',
            'Onay Durumu', 'Musteri ID', 'Musteri Ad Soyad', 'Telefon',
            'RFM Segment', 'Musteri Tipi',
            'Onerilen Urun 1', 'Onerilen Urun 2', 'Onerilen Urun 3', 'Onerilen Urun 4', 'Onerilen Urun 5'
        ]
        col_widths = [16, 20, 20, 14, 12, 24, 16, 22, 14, 26, 26, 26, 26, 26]
        for ci, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font = font(bold=True, color=C_WHITE, size=10)
            c.fill = fill(C_PURPLE_MID)
            c.alignment = align('center')
            c.border = border_all(C_PURPLE_MID)
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[2].height = 22

        data_row = 3
        row_fills = [C_WHITE, C_GREY_LIGHT]

        def write_rows2(camp, customers):
            nonlocal data_row
            u = _parse_onerilen_urunler(camp)
            camp_name = db_engine.val(camp, 'kampanya_tipi', '') or ''
            kaynak    = db_engine.val(camp, 'kaynak_kategori_ad') or ''
            hedef     = db_engine.val(camp, 'kategori_ad') or ''
            for idx, c in enumerate(customers):
                row_fill = row_fills[idx % 2]
                values = [
                    camp_name, kaynak, hedef,
                    db_engine.val(c, 'onay_durumu', 'Bilinmiyor'),
                    db_engine.val(c, 'id', ''),
                    db_engine.val(c, 'ad', ''),
                    db_engine.val(c, 'telefon', ''),
                    db_engine.val(c, 'rfm_segment', ''),
                    db_engine.val(c, 'tip', ''),
                    u[0] if len(u) > 0 else '',
                    u[1] if len(u) > 1 else '',
                    u[2] if len(u) > 2 else '',
                    u[3] if len(u) > 3 else '',
                    u[4] if len(u) > 4 else '',
                ]
                for ci, val in enumerate(values, start=1):
                    cell = ws2.cell(row=data_row, column=ci, value=val)
                    cell.font = font(size=9, color=C_TEXT_DARK)
                    cell.fill = fill(row_fill)
                    cell.alignment = align('left')
                    cell.border = border_bottom('E2E8F0')
                ws2.row_dimensions[data_row].height = 18
                data_row += 1

        for camp, source_cat_id, target_cat_id in cross_sell_camps:
            write_rows2(camp, customer_map.get((source_cat_id, target_cat_id), []))
        for camp, segment_clean in segment_camps:
            write_rows2(camp, segment_customer_map.get(segment_clean, []))

        logger.info(f"[EXPORT] adım5 excel_rows_yazildi t={_time.time()-_t0:.2f}s")
        # ── Kaydet ──────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        logger.info(f"[EXPORT] adım6 wb_kaydedildi t={_time.time()-_t0:.2f}s")

        camp_types = list({db_engine.val(c, 'kampanya_tipi', '') for c in camps if db_engine.val(c, 'kampanya_tipi')})
        type_label = '-'.join(sorted(camp_types)) if camp_types else 'Genel'
        filename = f"Kampanya_Raporu_{type_label}_{len(oneri_ids)}_Kampanya_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_data = output.read()
        logger.info(f"[EXPORT] tamamlandi boyut={len(file_data)} t={_time.time()-_t0:.2f}s")
        # StreamingHttpResponse: Railway 60s proxy timeout'unu aşmak için
        response = HttpResponse(
            file_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Length'] = len(file_data)
        return response

    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return HttpResponse(f"Hata: {str(e)}", status=400)
    finally:
        db_engine.release_connection(conn)


@csrf_exempt
def export_campaign_customers_excel(request):
    if request.method == 'GET':
        ids_str = request.GET.get('ids', '')
        if not ids_str:
            return HttpResponse('ids parametresi gerekli', status=400)
        try:
            oneri_ids = [int(x.strip()) for x in ids_str.split(',') if x.strip()]
        except ValueError:
            return HttpResponse('Geçersiz ID formatı', status=400)
        return _build_excel_response(oneri_ids)

    if request.method != 'POST':
        return HttpResponse('Method not allowed', status=405)

    oneri_ids = []
    content_type = request.content_type or ''
    if 'application/json' in content_type:
        try:
            data = json.loads(request.body)
            oneri_ids = data.get('ids', [])
        except (json.JSONDecodeError, Exception):
            pass
    else:
        ids_json = request.POST.get('ids_json', '[]')
        try:
            oneri_ids = json.loads(ids_json)
        except (json.JSONDecodeError, Exception):
            pass

    if not oneri_ids:
        return HttpResponse('Lütfen en az bir kampanya seçin', status=400)

    return _build_excel_response(oneri_ids)


def get_campaign_target_customers(oneri_id, limit=2000):
    """Tek kampanya için hedef müşteri listesi — SMS/E-mail entegrasyonu için."""
    conn = db_engine.get_connection()
    cursor = db_engine.get_dict_cursor(conn)
    ph = db_engine.ph()

    cursor.execute(f"SELECT * FROM otomatikkampanyaonerileri WHERE oneri_id = {ph}", [oneri_id])
    r = cursor.fetchone()
    if not r:
        db_engine.release_connection(conn)
        return []

    oneri_tipi    = db_engine.val(r, 'kampanya_tipi', '')
    target_cat_id = db_engine.val(r, 'kategori_id')
    source_cat_id = db_engine.val(r, 'ikinci_urun_id')
    source_cat_ad = db_engine.val(r, 'ikinci_urun_ad') or db_engine.val(r, 'kaynak_kategori_ad', '')
    hedef_segment = db_engine.val(r, 'hedef_segment', '')

    if oneri_tipi == 'Cross-Sell' and not source_cat_id and source_cat_ad:
        try:
            for cat_n in [c.strip() for c in source_cat_ad.split('&')]:
                cursor.execute(
                    f"SELECT id FROM kategoriler WHERE (alt2 = {ph} OR alt1 = {ph} OR ana = {ph}) LIMIT 1",
                    [cat_n, cat_n, cat_n]
                )
                row = cursor.fetchone()
                if row and row['id'] != target_cat_id:
                    source_cat_id = row['id']
                    break
        except Exception as e:
            logger.warning(f"Category recovery error: {e}")

    customers = []
    try:
        if oneri_tipi == 'Cross-Sell' and source_cat_id:
            cursor.execute(f"""
                SELECT m.id, m.ad, m.telefon, m.rfm_segment, m.tip, m.onay_durumu
                FROM musteriler m
                WHERE m.id IN (
                    SELECT DISTINCT musteri_id FROM satislar
                    WHERE kategori_id = {ph} AND musteri_id IS NOT NULL
                    EXCEPT
                    SELECT DISTINCT musteri_id FROM satislar
                    WHERE kategori_id = {ph} AND musteri_id IS NOT NULL
                )
                LIMIT {ph}
            """, [source_cat_id, target_cat_id, limit])
            customers = cursor.fetchall()
        elif oneri_tipi in ['Loyalty', 'Win-Back', 'Retention'] or 'Alıcıları' in hedef_segment:
            segment_clean = hedef_segment.replace(' Alıcıları', '').strip()
            cursor.execute(
                f"SELECT id, ad, telefon, rfm_segment, tip, onay_durumu FROM musteriler WHERE rfm_segment LIKE {ph} LIMIT {ph}",
                [f"%{segment_clean}%", limit]
            )
            customers = cursor.fetchall()
    except Exception as e:
        logger.error(f"Target customer fetch error: {e}")

    db_engine.release_connection(conn)
    return [dict(c) if not isinstance(c, dict) else c for c in customers]


def clean_xml_string(value):
    """Excel (openpyxl/XML) tarafında hata veren kontrol karakterlerini temizler."""
    if not isinstance(value, str):
        return value
    # XML 1.0 standardına göre geçersiz karakterler: \x00-\x08, \x0b-\x0c, \x0e-\x1f
    return "".join(ch for ch in value if ch == "\t" or ch == "\n" or ch == "\r" or ord(ch) >= 32)

@csrf_exempt
def export_brand_customers(request):
    """Marka + Kategori bazlı müşteri listesini Excel olarak dışa aktar."""
    from datetime import datetime
    import time as _time
    import logging
    import io
    from django.http import HttpResponse, StreamingHttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    logger = logging.getLogger(__name__)
    _t0 = _time.time()
    
    brand = request.GET.get('brand', '')
    category = request.GET.get('category', '')
    level = request.GET.get('level', 'ana')  # ana, alt1, alt2

    logger.info(f"[BRAND_EXPORT] Başladı: brand='{brand}' category='{category}' level='{level}'")

    if not brand:
        return HttpResponse('brand parametresi gerekli', status=400)
    if not category:
        return HttpResponse('category parametresi gerekli', status=400)
    if level not in ('ana', 'alt1', 'alt2'):
        level = 'ana'

    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()

        # 1. Marka ID'lerini bul
        cursor.execute(f"SELECT id FROM markalar WHERE ad = {ph}", [brand])
        brand_ids = [db_engine.val(r, 'id') for r in cursor.fetchall()]
        if not brand_ids:
            logger.warning(f"[BRAND_EXPORT] Marka bulunamadı: {brand}")
            return HttpResponse(f'Marka bulunamadı: {brand}', status=404)

        # 2. Kategori ID'lerini bul
        cursor.execute(f"SELECT id FROM kategoriler WHERE {level} = {ph}", [category])
        cat_ids = [db_engine.val(r, 'id') for r in cursor.fetchall()]
        
        # Robustness: Seviye eşleşmezse diğer kolonlarda da ara
        if not cat_ids:
            logger.info(f"[BRAND_EXPORT] Kategori '{category}' {level} seviyesinde bulunamadı, diğer seviyeler deneniyor...")
            for l in ['alt1', 'alt2', 'ana']:
                if l == level: continue
                cursor.execute(f"SELECT id FROM kategoriler WHERE {l} = {ph}", [category])
                c_rows = cursor.fetchall()
                if c_rows:
                    cat_ids = [db_engine.val(r, 'id') for r in c_rows]
                    logger.info(f"[BRAND_EXPORT] Kategori '{category}' {l} seviyesinde bulundu.")
                    break

        if not cat_ids:
            logger.warning(f"[BRAND_EXPORT] Kategori bulunamadı: {category}")
            return HttpResponse(f'Kategori bulunamadı: {category}', status=404)

        # 3. Bu marka+kategorideki müşterileri çek
        b_ph = ','.join([ph] * len(brand_ids))
        c_ph = ','.join([ph] * len(cat_ids))

        cursor.execute(f"""
            SELECT 
                s.musteri_id,
                SUM(s.tutar) as harcama,
                COUNT(DISTINCT s.fis_no) as fis_adet,
                MAX(s.tarih) as son_t,
                MIN(s.tarih) as ilk_t
            FROM satislar s
            WHERE s.marka_id IN ({b_ph})
              AND s.kategori_id IN ({c_ph})
              AND s.musteri_id IS NOT NULL
            GROUP BY s.musteri_id
            ORDER BY harcama DESC
            LIMIT 5000
        """, brand_ids + cat_ids)

        agg_rows = cursor.fetchall()
        logger.info(f"[BRAND_EXPORT] {len(agg_rows)} agrega müşteri bulundu t={_time.time()-_t0:.2f}s")
        
        if not agg_rows:
            return HttpResponse(f'Bu seçim için müşteri bulunamadı ({brand} - {category})', status=404)

        # 4. Müşteri detaylarını çek
        m_ids = [db_engine.val(r, 'musteri_id') for r in agg_rows]
        m_ph = ','.join([ph] * len(m_ids))
        cursor.execute(f"SELECT id, ad, telefon, onay_durumu, rfm_segment, tip FROM musteriler WHERE id IN ({m_ph})", m_ids)
        m_rows_raw = cursor.fetchall()
        # id: row eşleşmesi (db_engine.val ile güvenli key erişimi)
        m_rows = {}
        for r in m_rows_raw:
            mid = db_engine.val(r, 'id')
            if mid:
                m_rows[mid] = r
        
        # 5. Birleştir
        customers = []
        for r in agg_rows:
            m_id = db_engine.val(r, 'musteri_id')
            m_info = m_rows.get(m_id, {})
            
            # Tarih formatlama (db_engine.val None dönebilir)
            son_t = db_engine.val(r, 'son_t', None)
            ilk_t = db_engine.val(r, 'ilk_t', None)
            
            customers.append({
                'id': m_id,
                'ad': clean_xml_string(db_engine.val(m_info, 'ad', str(m_id))),
                'telefon': clean_xml_string(db_engine.val(m_info, 'telefon', '')),
                'onay_durumu': clean_xml_string(db_engine.val(m_info, 'onay_durumu', '')),
                'rfm_segment': clean_xml_string(db_engine.val(m_info, 'rfm_segment', '')),
                'tip': clean_xml_string(db_engine.val(m_info, 'tip', '')),
                'toplam_harcama': db_engine.val(r, 'harcama', 0),
                'alisveris_sayisi': db_engine.val(r, 'fis_adet', 0),
                'son_alisveris': son_t,
                'ilk_alisveris': ilk_t,
            })
            
        logger.info(f"[BRAND_EXPORT] {len(customers)} müşteri detaylandırıldı t={_time.time()-_t0:.2f}s")

        if not customers:
            return HttpResponse(f'Detaylandırma sonrası müşteri kalmadı ({brand} - {category})', status=404)

        # 6. Excel oluştur
        C_PURPLE_DARK = '3730A3'
        C_PURPLE_MID = '6366F1'
        C_WHITE = 'FFFFFF'
        C_GREY_LIGHT = 'F8FAFC'
        C_GREY_MID = 'E2E8F0'
        C_TEXT_DARK = '1E293B'

        # Module-level helpers (defined at top of file)
        fill = _xl_fill
        font = _xl_font
        border_bottom = _xl_border_bottom
        border_all = _xl_border_all
        align = _xl_align

        wb = Workbook()
        ws = wb.active
        ws.title = 'Müşteri Listesi'
        ws.sheet_view.showGridLines = False

        # Başlık banner
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = clean_xml_string(f'{brand} — {category} MÜŞTERİ LİSTESİ')
        title_cell.font = Font(name='Calibri', bold=True, color=C_WHITE, size=14)
        title_cell.fill = fill(C_PURPLE_DARK)
        title_cell.alignment = align('center')
        ws.row_dimensions[1].height = 36

        ws.merge_cells('A2:H2')
        sub_cell = ws['A2']
        sub_cell.value = clean_xml_string(f'Toplam {len(customers)} müşteri  |  Oluşturulma: {datetime.now().strftime("%d.%m.%Y %H:%M")}')
        sub_cell.font = font(color='A5B4FC', size=10)
        sub_cell.fill = fill(C_PURPLE_DARK)
        sub_cell.alignment = align('center')
        ws.row_dimensions[2].height = 22

        # Kolon başlıkları
        col_headers = ['Müşteri ID', 'Ad Soyad', 'Telefon', 'Onay Durumu', 'RFM Segment', 'Toplam Harcama', 'Alışveriş Sayısı', 'Son Alışveriş']
        col_widths = [12, 28, 18, 14, 24, 18, 16, 16]

        for ci, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = font(bold=True, color=C_WHITE, size=10)
            c.fill = fill(C_PURPLE_MID)
            c.alignment = align('center')
            c.border = border_all(C_PURPLE_MID)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[3].height = 24

        # Veri satırları
        row_fills = [C_WHITE, C_GREY_LIGHT]
        for idx, cust in enumerate(customers):
            row_idx = idx + 4
            row_fill = row_fills[idx % 2]
            
            harcama = float(cust.get('toplam_harcama', 0) or 0)
            son_tarih = cust.get('son_alisveris', '')
            if hasattr(son_tarih, 'strftime'):
                son_tarih_str = son_tarih.strftime('%d.%m.%Y')
            else:
                son_tarih_str = str(son_tarih)[:10] if son_tarih else ''

            values = [
                cust.get('id', ''),
                cust.get('ad', ''),
                cust.get('telefon', ''),
                cust.get('onay_durumu', ''),
                cust.get('rfm_segment', ''),
                harcama,
                int(cust.get('alisveris_sayisi', 0) or 0),
                son_tarih_str
            ]

            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font = font(size=10)
                cell.fill = fill(row_fill)
                cell.alignment = align('left' if ci <= 5 else 'right')
                cell.border = border_bottom()
                if ci == 6 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
            ws.row_dimensions[row_idx].height = 20

        # Kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = output.read()

        # Filename normalization (ASCII-safe)
        safe_brand = db_engine.normalize_turkish_py(brand).replace(' ', '_')[:20]
        safe_cat = db_engine.normalize_turkish_py(category).replace(' ', '_')[:20]
        filename = f"Musteri_Listesi_{safe_brand}_{safe_cat}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        logger.info(f"[BRAND_EXPORT] Tamamlandı: boyut={len(file_data)} t={_time.time()-_t0:.2f}s")

        from urllib.parse import quote
        encoded_filename = quote(filename)
        
        response = HttpResponse(
            file_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Use filename* for better UTF-8 support and browser compatibility
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Length'] = len(file_data)
        return response

    except Exception as e:
        logger.error(f"[BRAND_EXPORT] KRİTİK HATA: {e}", exc_info=True)
        return HttpResponse(f"Dışa aktarma sırasında bir hata oluştu: {str(e)}", status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)


@csrf_exempt
def export_beklenen_musteriler_excel(request):
    """Bu hafta beklenen veya geciken müşteri listesini Excel olarak dışa aktar."""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote

    tip = request.GET.get('tip', 'beklenen')
    magaza_id = request.GET.get('magaza_id', '').strip()
    conn = None

    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)

        tahmini = db_engine.col_date_add_expr('v.son_ziyaret_tarihi', 'v.ort_ziyaret_araligi')
        today = "CURRENT_DATE" if db_engine.DB_BACKEND == 'postgresql' else "date('now')"
        last90 = db_engine.date_offset_expr(-90)
        week_start = db_engine.date_trunc_expr('week', today)
        yesterday = db_engine.date_offset_expr(-1)
        week_end = db_engine.date_offset_expr(6, week_start)
        last_sunday = db_engine.date_offset_expr(-1, week_start)

        if tip == 'geciken':
            date_filter = f"{tahmini} BETWEEN {last90} AND {last_sunday}"
            order_sql = "tahmini_ziyaret_tarihi DESC"
            sheet_title = 'Geciken Müşteriler'
            banner_title = 'GECİKEN MÜŞTERİLER — Tahmini Ziyaret Tarihi Geçmiş'
        else:
            date_filter = f"{tahmini} BETWEEN {week_start} AND {week_end}"
            order_sql = "tahmini_ziyaret_tarihi ASC"
            sheet_title = 'Bu Hafta Beklenen'
            banner_title = 'BU HAFTA BEKLENEN MÜŞTERİLER'
        
        magaza_filter = ""
        magaza_params = []
        if magaza_id:
            magaza_filter = f" AND m.kayit_magazasi = {db_engine.ph()}"
            magaza_params.append(magaza_id)

        cursor.execute(f"""
            SELECT
                v.musteri_id,
                o.ad_soyad,
                m.telefon,
                o.rfm_segment,
                v.son_ziyaret_tarihi,
                {tahmini} AS tahmini_ziyaret_tarihi,
                ROUND(CAST(v.ort_ziyaret_araligi AS REAL), 1) AS ort_aralik_gun,
                CAST(v.toplam_ziyaret AS INTEGER) AS toplam_ziyaret,
                {db_engine.date_diff_days_expr(today, tahmini)} AS gecikme_gun,
                COALESCE(o.toplam_harcama, 0) AS toplam_harcama,
                COALESCE(o.ortalama_sepet_tutari, 0) AS ortalama_sepet_tutari,
                COALESCE(o.ortalama_sepet_tutari, 0) AS tahmini_alisveris_tutari,
                CASE
                    WHEN v.std_ziyaret_araligi / NULLIF(v.ort_ziyaret_araligi, 0) < 0.3
                         AND v.toplam_ziyaret >= 10 THEN 'Yuksek'
                    WHEN v.std_ziyaret_araligi / NULLIF(v.ort_ziyaret_araligi, 0) < 0.5
                         AND v.toplam_ziyaret >= 5  THEN 'Orta'
                    ELSE 'Dusuk'
                END AS guven_skoru,
                mg.ad as magaza_adi
            FROM musteriziyaretfeatures v
            JOIN musteriler m ON v.musteri_id = m.id
            JOIN musteridetayozet o ON v.musteri_id = o.musteri_id
            LEFT JOIN musterietiketler e ON v.musteri_id = e.musteri_id
            LEFT JOIN magazalar mg ON m.kayit_magazasi = mg.id::text
            WHERE v.ort_ziyaret_araligi IS NOT NULL
              AND v.toplam_ziyaret >= 3
              AND COALESCE(e.tamamen_kaybedilmis, FALSE) = FALSE
              AND {date_filter} {magaza_filter}
            ORDER BY {order_sql}
        """, magaza_params)
        rows = cursor.fetchall()
        db_engine.release_connection(conn)
        conn = None

        # Excel oluştur
        C_AMBER = 'B45309' if tip == 'beklenen' else '991B1B'
        C_AMBER_LIGHT = 'F59E0B' if tip == 'beklenen' else 'EF4444'
        C_WHITE = 'FFFFFF'
        C_GREY_LIGHT = 'F8FAFC'
        C_GREY_MID = 'E2E8F0'
        C_TEXT_DARK = '1E293B'

        # Module-level helpers (defined at top of file)
        _fill = _xl_fill
        _font = _xl_font
        _align = _xl_align
        _border = lambda: _xl_border_all()

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.sheet_view.showGridLines = False

        col_headers = ['Müşteri ID', 'Ad Soyad', 'Telefon', 'Son Ziyaret', 'Tahmini Ziyaret',
                        'Gecikme (Gün)', 'Ort. Aralık (Gün)', 'Ziyaret Sayısı', 'RFM Segment', 'Güven',
                        'Toplam Harcama (₺)', 'Ort. Sepet (₺)', 'Tahmini Tutar (₺)']
        col_widths = [12, 28, 18, 14, 14, 14, 16, 14, 24, 12, 18, 16, 18]

        # Banner
        ws.merge_cells(f'A1:{get_column_letter(len(col_headers))}1')
        title_cell = ws['A1']
        title_cell.value = clean_xml_string(banner_title)
        title_cell.font = Font(name='Calibri', bold=True, color=C_WHITE, size=14)
        title_cell.fill = _fill(C_AMBER)
        title_cell.alignment = _align('center')
        ws.row_dimensions[1].height = 36

        ws.merge_cells(f'A2:{get_column_letter(len(col_headers))}2')
        sub_cell = ws['A2']
        sub_cell.value = clean_xml_string(f'Toplam {len(rows)} müşteri  |  Oluşturulma: {datetime.now().strftime("%d.%m.%Y %H:%M")}')
        sub_cell.font = _font(color=C_WHITE, size=10)
        sub_cell.fill = _fill(C_AMBER)
        sub_cell.alignment = _align('center')
        ws.row_dimensions[2].height = 22

        # Kolon başlıkları
        for ci, (header, width) in enumerate(zip(col_headers, col_widths), 1):
            cell = ws.cell(row=3, column=ci, value=header)
            cell.font = _font(bold=True, color=C_WHITE, size=10)
            cell.fill = _fill(C_AMBER_LIGHT)
            cell.alignment = _align('center')
            cell.border = _border()
            ws.column_dimensions[get_column_letter(ci)].width = width

        # Veriler
        for ri, row in enumerate(rows, 4):
            bg = C_GREY_LIGHT if ri % 2 == 0 else C_WHITE
            vals = [
                db_engine.val(row, 'musteri_id'),
                clean_xml_string(str(db_engine.val(row, 'ad_soyad', '') or '')),
                clean_xml_string(str(db_engine.val(row, 'telefon', '') or '')),
                str(db_engine.val(row, 'son_ziyaret_tarihi', '') or '')[:10],
                str(db_engine.val(row, 'tahmini_ziyaret_tarihi', '') or '')[:10],
                db_engine.val(row, 'gecikme_gun', 0),
                db_engine.val(row, 'ort_aralik_gun', 0),
                db_engine.val(row, 'toplam_ziyaret', 0),
                clean_xml_string(str(db_engine.val(row, 'rfm_segment', '') or '')),
                db_engine.val(row, 'guven_skoru', ''),
                float(db_engine.val(row, 'toplam_harcama', 0) or 0),
                float(db_engine.val(row, 'ortalama_sepet_tutari', 0) or 0),
                float(db_engine.val(row, 'tahmini_alisveris_tutari', 0) or 0),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = _font(size=10)
                cell.fill = _fill(bg)
                cell.alignment = _align('center') if ci >= 4 else _align()
                cell.border = _border()
                if ci >= 11 and isinstance(v, (int, float)):
                    cell.number_format = '#,##0.00'

        # Response
        buffer = io.BytesIO()
        wb.save(buffer)
        file_data = buffer.getvalue()

        tip_label = 'Geciken' if tip == 'geciken' else 'Bu_Hafta_Beklenen'
        fname = f'{tip_label}_Musteriler_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        encoded = quote(fname)

        response = HttpResponse(file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Length'] = len(file_data)
        return response

    except Exception as e:
        logger.error(f"[BEKLENEN_EXPORT] HATA: {e}", exc_info=True)
        return HttpResponse(f"Dışa aktarma hatası: {str(e)}", status=500)
    finally:
        if conn:
            db_engine.release_connection(conn)
