"""
Comprehensive CRM Category Analytics & Tagging
"""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST, HTTP_401_UNAUTHORIZED, HTTP_500_INTERNAL_SERVER_ERROR
import logging
import time as _time
from datetime import datetime
from collections import defaultdict

from .. import db_engine
from .base import get_user_from_request

logger = logging.getLogger(__name__)

# Performans için kritik indeksler
_cat_indexes_created = False

def _ensure_category_indexes(cursor):
    global _cat_indexes_created
    if _cat_indexes_created:
        return
    if db_engine.DB_BACKEND != 'postgresql':
        _cat_indexes_created = True
        return
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_satislar_kategori_id ON satislar (kategori_id)",
        "CREATE INDEX IF NOT EXISTS idx_satislar_kategori_tutar ON satislar (kategori_id, tutar)",
        "CREATE INDEX IF NOT EXISTS idx_satislar_kategori_marka ON satislar (kategori_id, marka_id)",
        "CREATE INDEX IF NOT EXISTS idx_satislar_kat_musteri ON satislar (kategori_id, musteri_id) WHERE musteri_id IS NOT NULL",
        "CREATE INDEX IF NOT EXISTS idx_kategoriler_ana ON kategoriler (ana)",
        "CREATE INDEX IF NOT EXISTS idx_kategoriler_alt1 ON kategoriler (alt1)",
        "CREATE INDEX IF NOT EXISTS idx_kategoriler_alt2 ON kategoriler (alt2)",
        "CREATE INDEX IF NOT EXISTS idx_grupbirliktelikleri_kat1 ON grupbirliktelikleri (kategori_id_1, tip)",
    ]
    conn = cursor.connection
    for idx_sql in indexes:
        try:
            # Her index'i ayrı savepoint içinde çalıştır — hata olursa sadece o geri alınır
            cursor.execute("SAVEPOINT idx_sp")
            cursor.execute(idx_sql)
            cursor.execute("RELEASE SAVEPOINT idx_sp")
        except Exception:
            try:
                cursor.execute("ROLLBACK TO SAVEPOINT idx_sp")
                cursor.execute("RELEASE SAVEPOINT idx_sp")
            except Exception:
                pass
    _cat_indexes_created = True

def get_db_connection():
    return db_engine.get_connection()

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_category_report_details(request, data_source_id):
    """
    Kategori bazlı kapsamlı CRM analizi. Hiyerarşik seviye bazlı çalışır.
    """
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz erişim'}, status=HTTP_401_UNAUTHORIZED)

    category_name = request.GET.get('name')
    level = request.GET.get('level', 'ana') # ana, alt1, alt2

    # Level parametresini validate et (SQL injection koruması)
    if level not in ('ana', 'alt1', 'alt2'):
        level = 'ana'

    if not category_name:
        return Response({'error': 'Kategori adı gerekli'}, status=HTTP_400_BAD_REQUEST)

    # Müşteri filtreleri
    customer_type = request.GET.get('customer_type') or request.GET.get('customerType')
    approval_status = request.GET.get('approval_status') or request.GET.get('approvalStatus')
    region = request.GET.get('region')
    has_customer_filter = bool(customer_type or approval_status or region)

    conn = None
    try:
        conn = get_db_connection()
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()

        # Index'leri bir kez oluştur
        _ensure_category_indexes(cursor)

        # Bu seviyedeki tüm kategori ID'lerini al
        cursor.execute(f"SELECT id FROM kategoriler WHERE {level} = {ph}", (category_name,))
        cat_ids = [row['id'] for row in cursor.fetchall()]

        if not cat_ids:
             return Response({'error': 'Kategori bulunamadı'}, status=HTTP_400_BAD_REQUEST)

        # Müşteri filtresi subquery'si
        musteri_filter_sql = ""
        musteri_filter_params = []
        if has_customer_filter:
            filter_parts = []
            if customer_type:
                filter_parts.append(f"tip = {ph}")
                musteri_filter_params.append(customer_type)
            if approval_status:
                filter_parts.append(f"onay_durumu = {ph}")
                musteri_filter_params.append(approval_status)
            if region:
                filter_parts.append(f"kayit_magazasi IN (SELECT id::text FROM magazalar WHERE bolge = {ph})")
                musteri_filter_params.append(region)
            where_str = " AND ".join(filter_parts)
            musteri_filter_sql = f"AND s.musteri_id IN (SELECT id FROM musteriler WHERE {where_str})"

        # ----------------------------------------------------------------
        # Önbellekten oku — yalnızca filtre yoksa
        # ----------------------------------------------------------------
        cached = None
        if not has_customer_filter:
            try:
                cursor.execute(
                    "SELECT * FROM kategori_analiz_ozet WHERE kategori_adi = %s AND level = %s",
                    (category_name, level),
                )
                cached = cursor.fetchone()
            except Exception:
                cached = None

        if cached:
            import json as _json
            cached = dict(cached)

            # Birliktelik analizi — hızlı, her zaman canlı hesaplanır
            strategy = request.GET.get('strategy', 'discovery')
            level_field_map = {'ana': 'ana', 'alt1': 'alt1', 'alt2': 'alt2'}
            group_field = level_field_map.get(level, 'ana')

            placeholders_live = ','.join([ph] * len(cat_ids))
            try:
                cursor.execute(f"""
                    SELECT
                        k2.{group_field} as category_name,
                        AVG(gb.confidence) as confidence,
                        AVG(gb.lift) as lift,
                        SUM(gb.ortak_fis_sayisi) as ortak_fis_sayisi,
                        SUM(gb.ortak_musteri_sayisi) as ortak_musteri_sayisi
                    FROM grupbirliktelikleri gb
                    JOIN kategoriler k2 ON gb.kategori_id_2 = k2.id
                    WHERE gb.kategori_id_1 IN ({placeholders_live})
                      AND k2.{group_field} != {ph}
                      AND k2.{group_field} IS NOT NULL
                      AND gb.tip = 'CAT_ONLY_SQL'
                    GROUP BY k2.{group_field}
                    ORDER BY lift DESC, ortak_fis_sayisi DESC
                    LIMIT 50
                """, (*cat_ids, category_name))
                raw_assoc = [dict(r) for r in cursor.fetchall()]

                # Birliktelik Analizi (Associations) - Artık cache tablosundan geliyor!
                try:
                    associations = _json.loads(cached.get('associations_json') or '[]')
                    # Sadece ilk 10 tanesini döndür (Frontend beklentisi)
                    associations = associations[:10]
                except Exception as assoc_err:
                    logger.warning(f"Cache path — associations error (non-fatal): {assoc_err}")
                    associations = []
            except Exception as assoc_err:
                logger.warning(f"Cache path — associations error (non-fatal): {assoc_err}")
                associations = []

            db_engine.release_connection(conn)

            return Response({
                'info': {'ana': category_name, 'level': level, 'etiketler': ''},
                'kpis': {
                    'total_revenue':   float(cached.get('total_revenue') or 0),
                    'total_receipts':  int(cached.get('total_receipts') or 0),
                    'total_customers': int(cached.get('total_customers') or 0),
                    'total_quantity':  float(cached.get('total_quantity') or 0),
                    'avg_price':       float(cached.get('avg_price') or 0),
                },
                'trends':               _json.loads(cached.get('trends_json') or '[]'),
                'topProducts':          _json.loads(cached.get('top_products_json') or '[]'),
                'rfmDistribution':      _json.loads(cached.get('rfm_json') or '[]'),
                'associations':         associations,
                'comparison':           _json.loads(cached.get('comparison_json') or '{}'),
                'brandTrends':          _json.loads(cached.get('brand_trends_json') or '[]'),
                'brandCustomerAnalysis': _json.loads(cached.get('brand_customer_json') or '[]'),
            }, status=HTTP_200_OK)
        # ----------------------------------------------------------------
        # Önbellekte yok — canlı hesapla (eski yol)
        # ----------------------------------------------------------------

        placeholders = ','.join([ph] * len(cat_ids))

        # 1. Kategori Bilgileri & Etiketler
        string_agg = "STRING_AGG(DISTINCT et.etiket, ',')" if db_engine.DB_BACKEND == 'postgresql' else "GROUP_CONCAT(DISTINCT et.etiket)"
        cursor.execute(f"""
            SELECT {ph} as ana, {ph} as level, {string_agg} as etiketler
            FROM kategorietiketleri et
            WHERE et.kategori_id IN ({placeholders})
        """, (category_name, level, *cat_ids))
        row = cursor.fetchone()
        category_info = dict(row) if row else {'ana': category_name, 'level': level, 'etiketler': ''}

        # 2. Genel Metrikler (KPIs)
        if has_customer_filter:
            cursor.execute(f"""
                SELECT
                    SUM(s.tutar) as total_revenue,
                    COUNT(DISTINCT s.fis_no) as total_receipts,
                    COUNT(DISTINCT s.musteri_id) as total_customers,
                    SUM(s.miktar) as total_quantity,
                    SUM(s.tutar)/NULLIF(SUM(s.miktar), 0) as avg_price
                FROM satislar s
                WHERE s.kategori_id IN ({placeholders})
                  {musteri_filter_sql}
            """, (*cat_ids, *musteri_filter_params))
        else:
            cursor.execute(f"""
                SELECT
                    SUM(revenue) as total_revenue,
                    SUM(receipt_count) as total_receipts,
                    SUM(customer_count) as total_customers,
                    SUM(unit_count) as total_quantity,
                    SUM(revenue)/NULLIF(SUM(unit_count), 0) as avg_price
                FROM daily_metrics_summary
                WHERE kategori_id IN ({placeholders})
            """, cat_ids)
        kpis = dict(cursor.fetchone())

        # 3. Aylık trend
        month_expr = db_engine.strftime_expr('%Y-%m', 'tarih')
        if has_customer_filter:
            cursor.execute(f"""
                SELECT {db_engine.strftime_expr('%Y-%m', 's.tarih')} as month,
                       SUM(s.tutar) as revenue, SUM(s.miktar) as quantity
                FROM satislar s
                WHERE s.kategori_id IN ({placeholders})
                  {musteri_filter_sql}
                GROUP BY month
                ORDER BY month DESC
                LIMIT 12
            """, (*cat_ids, *musteri_filter_params))
        else:
            cursor.execute(f"""
                SELECT {month_expr} as month, SUM(revenue) as revenue, SUM(unit_count) as quantity
                FROM daily_metrics_summary
                WHERE kategori_id IN ({placeholders})
                GROUP BY month
                ORDER BY month DESC
                LIMIT 12
            """, cat_ids)
        trends = [dict(row) for row in cursor.fetchall()]

        # 4. Kategori İçi En Çok Satan Ürünler
        if has_customer_filter:
            cursor.execute(f"""
                SELECT
                    u.id as product_id,
                    u.ad as product_name,
                    SUM(s.tutar) as revenue,
                    SUM(s.miktar) as quantity,
                    COUNT(DISTINCT s.musteri_id) as customer_count,
                    COUNT(DISTINCT s.fis_no) as receipt_count
                FROM satislar s
                JOIN urunler u ON s.urun_id = u.id
                WHERE u.kategori_id IN ({placeholders})
                  {musteri_filter_sql}
                GROUP BY u.id, u.ad
                ORDER BY revenue DESC
                LIMIT 10
            """, (*cat_ids, *musteri_filter_params))
        else:
            cursor.execute(f"""
                SELECT
                    u.id as product_id,
                    u.ad as product_name,
                    SUM(ps.revenue) as revenue,
                    SUM(ps.unit_count) as quantity,
                    SUM(ps.customer_count) as customer_count,
                    SUM(ps.receipt_count) as receipt_count
                FROM product_daily_summary ps
                JOIN urunler u ON ps.urun_id = u.id
                WHERE u.kategori_id IN ({placeholders})
                GROUP BY u.id, u.ad
                ORDER BY revenue DESC
                LIMIT 10
            """, cat_ids)
        top_products = [dict(row) for row in cursor.fetchall()]

        # 5. RFM Segment Dağılımı
        if has_customer_filter:
            cursor.execute(f"""
                SELECT mu.rfm_segment as segment, COUNT(DISTINCT s.musteri_id) as count
                FROM satislar s
                JOIN musteriler mu ON s.musteri_id = mu.id
                WHERE s.kategori_id IN ({placeholders}) AND mu.rfm_segment IS NOT NULL
                  {musteri_filter_sql}
                GROUP BY mu.rfm_segment
            """, (*cat_ids, *musteri_filter_params))
        else:
            cursor.execute(f"""
                SELECT rfm_segment as segment, SUM(customer_count) as count
                FROM daily_metrics_summary
                WHERE kategori_id IN ({placeholders}) AND rfm_segment IS NOT NULL
                GROUP BY rfm_segment
            """, cat_ids)
        rfm_distribution = [dict(row) for row in cursor.fetchall()]

        # 6. Birliktelik Analizi (Fallback Path - Canlı Sorgu)
        level_field_map = {'ana': 'ana', 'alt1': 'alt1', 'alt2': 'alt2'}
        group_field = level_field_map.get(level, 'ana')
        
        try:
            # Önce birliktelik verilerini çek (join'i küçültmek için limitli)
            cursor.execute(f"""
                SELECT
                    k2.{group_field} as category_name,
                    AVG(gb.confidence) as confidence,
                    AVG(gb.lift) as lift,
                    SUM(gb.ortak_fis_sayisi) as ortak_fis_sayisi
                FROM grupbirliktelikleri gb
                JOIN kategoriler k2 ON gb.kategori_id_2 = k2.id
                WHERE gb.kategori_id_1 IN ({placeholders})
                  AND k2.{group_field} != {ph}
                  AND k2.{group_field} IS NOT NULL
                  AND gb.tip = 'CAT_ONLY_SQL'
                GROUP BY k2.{group_field}
                ORDER BY lift DESC
                LIMIT 10
            """, (*cat_ids, category_name))
            associations = [dict(row) for row in cursor.fetchall()]
        except Exception as assoc_err:
            logger.error(f"Fallback path — associations error: {assoc_err}")
            associations = []
            if db_engine.DB_BACKEND == 'postgresql':
                try:
                    conn.rollback()
                except Exception:
                    pass

        # 7. Karşılaştırmalı Analiz (Benchmarking)
        comparison = {
            'marketShare': 0,
            'parentName': None,
            'levelLabel': 'Ana Kategori' if level == 'ana' else ('Alt Kategori 1' if level == 'alt1' else 'Alt Kategori 2'),
            'benchmarks': {},
            'siblings': []
        }

        try:
            cursor.execute(f"SELECT ana, alt1, alt2 FROM kategoriler WHERE id = {ph}", (cat_ids[0],))
            hierarchy_row = cursor.fetchone()
            if hierarchy_row:
                hierarchy = dict(hierarchy_row)

                parent_level = None
                parent_name = None
                if level == 'alt2':
                    parent_level = 'alt1'
                    parent_name = hierarchy.get('alt1')
                elif level == 'alt1':
                    parent_level = 'ana'
                    parent_name = hierarchy.get('ana')

                if parent_name and parent_level:
                    cursor.execute(f"SELECT id FROM kategoriler WHERE {parent_level} = {ph}", (parent_name,))
                    p_cat_ids = [r['id'] for r in cursor.fetchall()]

                    if p_cat_ids:
                        p_placeholders = ','.join([ph] * len(p_cat_ids))

                        cursor.execute(f"""
                            SELECT
                                SUM(s.tutar) as rev, COUNT(DISTINCT s.musteri_id) as cust,
                                SUM(s.tutar)/NULLIF(SUM(s.miktar), 0) as avg_price
                            FROM satislar s
                            WHERE s.kategori_id IN ({p_placeholders})
                              {musteri_filter_sql}
                        """, (*p_cat_ids, *musteri_filter_params))
                        p_row = cursor.fetchone()
                        if p_row:
                            p_metrics = dict(p_row)
                            if p_metrics.get('rev'):
                                comparison['marketShare'] = (kpis['total_revenue'] / p_metrics['rev']) * 100
                                comparison['parentName'] = parent_name
                                comparison['benchmarks'] = {
                                    'parentRevenue': p_metrics['rev'],
                                    'parentAvgPrice': p_metrics.get('avg_price')
                                }

                        cursor.execute(f"""
                            SELECT k.{level} as name, SUM(s.tutar) as revenue
                            FROM satislar s
                            JOIN kategoriler k ON s.kategori_id = k.id
                            WHERE k.{parent_level} = {ph} AND k.{level} IS NOT NULL
                              {musteri_filter_sql}
                            GROUP BY k.{level}
                            ORDER BY revenue DESC
                            LIMIT 6
                        """, (parent_name, *musteri_filter_params))
                        comparison['siblings'] = [dict(row) for row in cursor.fetchall()]
        except Exception as comp_err:
            logger.warning(f"Comparison analysis error (non-fatal): {comp_err}")
            if db_engine.DB_BACKEND == 'postgresql':
                try:
                    conn.rollback()
                except Exception:
                    pass

        # 8. Marka Analizi (Pazar Payı Trendi)
        month_expr_brand = db_engine.strftime_expr('%Y-%m', 's.tarih')
        cursor.execute(f"""
            SELECT m.ad as name, {month_expr_brand} as month, SUM(s.tutar) as b_rev
            FROM satislar s
            JOIN markalar m ON s.marka_id = m.id
            WHERE s.kategori_id IN ({placeholders})
              {musteri_filter_sql}
            GROUP BY name, month
            ORDER BY month DESC
        """, (*cat_ids, *musteri_filter_params))

        raw_trends = [dict(row) for row in cursor.fetchall()]
        brand_map = defaultdict(list)
        for row in raw_trends:
            brand_map[row['name']].append({'month': row['month'], 'b_rev': row['b_rev']})

        cat_monthly_rev = {t['month']: t['revenue'] for t in trends}
        brand_trends = []
        for b_name, data in brand_map.items():
            shares = []
            total_b_rev = 0
            for r in data:
                m = r['month']
                b_rev = r['b_rev']
                total_b_rev += b_rev
                c_rev = cat_monthly_rev.get(m, 0)
                share = (b_rev / c_rev * 100) if c_rev > 0 else 0
                shares.append({'month': m, 'share': round(share, 2)})

            brand_trends.append({'name': b_name, 'data': shares, 'total_rev': total_b_rev})

        brand_trends.sort(key=lambda x: x['total_rev'], reverse=True)
        brand_trends = brand_trends[:50]

        # 9. Marka Müşteri Analizi (Penetrasyon)
        cursor.execute(f"""
            SELECT m.ad as name, COUNT(DISTINCT s.musteri_id) as customer_count
            FROM satislar s
            JOIN markalar m ON s.marka_id = m.id
            WHERE s.kategori_id IN ({placeholders})
              {musteri_filter_sql}
            GROUP BY name
            ORDER BY customer_count DESC
            LIMIT 50
        """, (*cat_ids, *musteri_filter_params))
        brand_customers_raw = [dict(row) for row in cursor.fetchall()]

        total_cat_customers = kpis.get('total_customers', 0)
        brand_customer_analysis = []
        for r in brand_customers_raw:
            c_count = r['customer_count']
            share = (c_count / total_cat_customers * 100) if total_cat_customers > 0 else 0
            brand_customer_analysis.append({
                'name': r['name'],
                'count': c_count,
                'share': round(share, 2)
            })

        db_engine.release_connection(conn)

        return Response({
            'info': category_info,
            'kpis': kpis,
            'trends': trends,
            'topProducts': top_products,
            'rfmDistribution': rfm_distribution,
            'associations': associations,
            'comparison': comparison,
            'brandTrends': brand_trends,
            'brandCustomerAnalysis': brand_customer_analysis
        }, status=HTTP_200_OK)

    except Exception as e:
        logger.error(f"Category report error: {e}")
        if conn:
            db_engine.release_connection(conn)
        return Response({'error': str(e)}, status=HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_category_report_excel(request, data_source_id):
    """Kategori analizini Excel olarak dışa aktar"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    category_name = request.GET.get('name')
    level = request.GET.get('level', 'ana')
    
    # Veriyi get_category_report_details mantığıyla çek
    # (Burada kod tekrarı yerine fonksiyonu çağırmak veya mantığı ortaklaştırmak daha iyi olabilir 
    # ama Response nesnesi döndüğü için manuel çekim daha temiz)
    
    conn = db_engine.get_connection()
    try:
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()
        
        # Level ve ID tespiti
        cursor.execute(f"SELECT id FROM kategoriler WHERE {level} = {ph}", (category_name,))
        cat_ids = [row['id'] for row in cursor.fetchall()]
        if not cat_ids:
            return HttpResponse("Kategori bulunamadı", status=400)
            
        placeholders = ','.join([ph] * len(cat_ids))
        
        # KPIs
        cursor.execute(f"""
            SELECT
                SUM(revenue) as total_revenue,
                SUM(receipt_count) as total_receipts,
                SUM(customer_count) as total_customers,
                SUM(unit_count) as total_quantity
            FROM daily_metrics_summary
            WHERE kategori_id IN ({placeholders})
        """, cat_ids)
        kpi = cursor.fetchone()
        
        # En Çok Satan Ürünler
        cursor.execute(f"""
            SELECT u.ad as product_name, SUM(ps.revenue) as revenue, SUM(ps.unit_count) as quantity
            FROM product_daily_summary ps
            JOIN urunler u ON ps.urun_id = u.id
            WHERE u.kategori_id IN ({placeholders})
            GROUP BY u.ad
            ORDER BY revenue DESC
            LIMIT 20
        """, cat_ids)
        top_products = cursor.fetchall()
        
        # Marka Analizi
        cursor.execute(f"""
            SELECT m.ad as name, SUM(ps.revenue) as revenue
            FROM product_daily_summary ps
            JOIN markalar m ON ps.marka_id = m.id
            WHERE ps.kategori_id IN ({placeholders})
            GROUP BY m.ad
            ORDER BY revenue DESC
            LIMIT 20
        """, cat_ids)
        brands = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kategori Analizi"
        
        # Başlık Bölümü
        ws.merge_cells('A1:C1')
        ws['A1'] = f"Kategori Analizi: {category_name} ({level})"
        ws['A1'].font = Font(size=14, bold=True)
        
        # KPI Tablosu
        ws.append([])
        ws.append(["Metrik", "Değer"])
        ws.append(["Toplam Ciro", kpi['total_revenue'] or 0])
        ws.append(["Fiş Adedi", kpi['total_receipts'] or 0])
        ws.append(["Müşteri Sayısı", kpi['total_customers'] or 0])
        ws.append(["Satılan Miktar", kpi['total_quantity'] or 0])
        
        # Stil
        for row in ws.iter_rows(min_row=3, max_row=7, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")
                if cell.row == 3: cell.font = Font(bold=True)

        # Ürün Tablosu
        ws.append([])
        ws.append(["En Çok Satan Ürünler", "Ciro", "Miktar"])
        for p in top_products:
            ws.append([p['product_name'], p['revenue'], p['quantity']])
            
        # Marka Tablosu
        ws.append([])
        ws.append(["Marka Dağılımı", "Ciro"])
        for b in brands:
            ws.append([b['name'], b['revenue']])

        # Güzelleştirme
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Kategori_{category_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Category Excel error: {e}")
        return HttpResponse(f"Hata: {str(e)}", status=500)
    finally:
        db_engine.release_connection(conn)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def handle_category_tags(request, data_source_id):
    """
    Kategoriye etiket ekle veya sil.
    """
    user = get_user_from_request(request)
    if not user:
        return Response({'error': 'Yetkisiz erişim'}, status=HTTP_401_UNAUTHORIZED)

    category_id = request.data.get('category_id')
    etiket = request.data.get('tag')
    action = request.data.get('action', 'add') # add or remove

    if not category_id or not etiket:
        return Response({'error': 'Kategori ID ve etiket gerekli'}, status=HTTP_400_BAD_REQUEST)

    try:
        conn = get_db_connection()
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()

        if action == 'add':
            cursor.execute(f"INSERT INTO kategorietiketleri (kategori_id, etiket) VALUES ({ph}, {ph})", (category_id, etiket))
        else:
            cursor.execute(f"DELETE FROM kategorietiketleri WHERE kategori_id = {ph} AND etiket = {ph}", (category_id, etiket))

        conn.commit()
        db_engine.release_connection(conn)

        return Response({'status': 'success'}, status=HTTP_200_OK)
    except Exception as e:
        logger.error(f"Tag handling error: {e}")
        return Response({'error': str(e)}, status=HTTP_500_INTERNAL_SERVER_ERROR)


_category_tree_cache = {}
_TREE_CACHE_TTL = 300  # 5 dakika

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_category_tree(request, data_source_id):
    """
    Tam hiyerarşik (ana -> alt1 -> alt2) kategori ağacını ciro verileriyle birlikte döner.
    Server-side 5 dk cache ile hızlandırılmıştır.
    """
    global _category_tree_cache
    cache_key = f"tree_{data_source_id}"
    now = _time.time()

    # Cache kontrol
    if cache_key in _category_tree_cache:
        cached_data, cached_time = _category_tree_cache[cache_key]
        if now - cached_time < _TREE_CACHE_TTL:
            return Response(cached_data, status=HTTP_200_OK)

    try:
        conn = get_db_connection()
        cursor = db_engine.get_dict_cursor(conn)

        cursor.execute("""
            SELECT k.ana, k.alt1, k.alt2, SUM(ms.revenue) as revenue
            FROM daily_metrics_summary ms
            JOIN kategoriler k ON ms.kategori_id = k.id
            GROUP BY k.ana, k.alt1, k.alt2
            ORDER BY k.ana, k.alt1, k.alt2
        """)

        rows = cursor.fetchall()

        tree = {}

        for row in rows:
            ana, alt1, alt2, rev = row['ana'], row['alt1'], row['alt2'], row['revenue']

            if ana not in tree:
                tree[ana] = {'name': ana, 'level': 'ana', 'revenue': 0, 'children': {}}

            tree[ana]['revenue'] += rev

            if alt1:
                if alt1 not in tree[ana]['children']:
                    tree[ana]['children'][alt1] = {'name': alt1, 'level': 'alt1', 'revenue': 0, 'children': {}}

                tree[ana]['children'][alt1]['revenue'] += rev

                if alt2:
                    if alt2 not in tree[ana]['children'][alt1]['children']:
                        tree[ana]['children'][alt1]['children'][alt2] = {'name': alt2, 'level': 'alt2', 'revenue': rev}

        def convert_to_list(node):
            if 'children' in node:
                children_list = []
                sorted_children = sorted(node['children'].values(), key=lambda x: x['revenue'], reverse=True)
                for child in sorted_children:
                    children_list.append(convert_to_list(child))
                node['children'] = children_list
            return node

        final_tree = []
        sorted_ana = sorted(tree.values(), key=lambda x: x['revenue'], reverse=True)
        for ana_node in sorted_ana:
            final_tree.append(convert_to_list(ana_node))

        db_engine.release_connection(conn)

        # Cache'e kaydet
        _category_tree_cache[cache_key] = (final_tree, now)

        return Response(final_tree, status=HTTP_200_OK)
    except Exception as e:
        logger.error(f"Category tree error: {e}")
        return Response({'error': str(e)}, status=HTTP_500_INTERNAL_SERVER_ERROR)

