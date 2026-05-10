import logging
from django.db import connection
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from .. import db_engine
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

logger = logging.getLogger(__name__)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_store_analysis(request):
    """Mağaza performans analizi ve karşılaştırma"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    region = request.GET.get('region')
    customer_type = request.GET.get('customer_type')
    approval_status = request.GET.get('approval_status')
    
    conn = db_engine.get_connection()
    try:
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()
        
        # Filtreleri hazırla
        where_parts = ["1=1"]
        params = []
        if start_date:
            where_parts.append(f"tarih >= {ph}")
            params.append(start_date)
        if end_date:
            where_parts.append(f"tarih <= {ph}")
            params.append(end_date)
        if region:
            where_parts.append(f"magaza_id IN (SELECT id FROM magazalar WHERE {db_engine.bolge_expr()} = {ph})")
            params.append(region)
        if customer_type:
            where_parts.append(f"LOWER(customer_type) = LOWER({ph})")
            params.append(customer_type)
        if approval_status:
            where_parts.append(f"LOWER(onay_durumu) = LOWER({ph})")
            params.append(approval_status)
            
        where_clause = " AND ".join(where_parts)
        
        # Mağaza bazlı özet veriler
        be = db_engine.bolge_expr('m.bolge')
        query = f"""
            SELECT 
                m.ad as magaza_ad,
                {be} as bolge,
                SUM(d.revenue) as ciro,
                SUM(d.receipt_count) as fis_adedi,
                SUM(d.unit_count) as miktar,
                COUNT(DISTINCT d.tarih) as gun_sayisi
            FROM daily_metrics_summary d
            JOIN magazalar m ON d.magaza_id = m.id
            WHERE {where_clause}
            GROUP BY m.ad, {be}
            ORDER BY ciro DESC
        """
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # Müşteri sayıları (Summary table'dan inflated gelebilir, bu yüzden ayrı bir sorgu gerekebilir)
        # Ancak performans için summary table SUM(customer_count) kullanılabilir veya 
        # satislar'dan çekilebilir. Burada mağaza bazlı olduğu için satislar'dan çekmek daha güvenli.
        
        store_stats = []
        for r in rows:
            ciro = float(r['ciro'] or 0)
            fis = int(r['fis_adedi'] or 0)
            
            store_stats.append({
                'magaza': r['magaza_ad'],
                'bolge': r['bolge'],
                'ciro': ciro,
                'fisAdedi': fis,
                'miktar': int(r['miktar'] or 0),
                'sepetOrtalaması': round(ciro / fis, 2) if fis > 0 else 0,
                'gunlukOrtCiro': round(ciro / r['gun_sayisi'], 2) if r['gun_sayisi'] > 0 else 0
            })
            
        return Response({
            'status': 'success',
            'data': store_stats,
            'summary': {
                'totalCiro': sum(s['ciro'] for s in store_stats),
                'totalFis': sum(s['fisAdedi'] for s in store_stats),
                'totalStores': len(store_stats)
            }
        })
    except Exception as e:
        logger.error(f"Store analysis error: {e}")
        return Response({'status': 'error', 'message': str(e)}, status=500)
    finally:
        db_engine.release_connection(conn)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_store_analysis_excel(request):
    """Mağaza analizini Excel formatında dışa aktar"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    region = request.GET.get('region')
    customer_type = request.GET.get('customer_type')
    approval_status = request.GET.get('approval_status')
    
    # Veriyi çek (get_store_analysis mantığıyla)
    conn = db_engine.get_connection()
    try:
        cursor = db_engine.get_dict_cursor(conn)
        ph = db_engine.ph()
        where_parts = ["1=1"]
        params = []
        if start_date:
            where_parts.append(f"tarih >= {ph}"); params.append(start_date)
        if end_date:
            where_parts.append(f"tarih <= {ph}"); params.append(end_date)
        if region:
            where_parts.append(f"magaza_id IN (SELECT id FROM magazalar WHERE {db_engine.bolge_expr()} = {ph})")
            params.append(region)
        if customer_type:
            where_parts.append(f"LOWER(customer_type) = LOWER({ph})")
            params.append(customer_type)
        if approval_status:
            where_parts.append(f"LOWER(onay_durumu) = LOWER({ph})")
            params.append(approval_status)
        
        where_clause = " AND ".join(where_parts)
        be = db_engine.bolge_expr('m.bolge')
        query = f"""
            SELECT 
                m.ad as magaza_ad, {be} as bolge,
                SUM(d.revenue) as ciro, SUM(d.receipt_count) as fis_adedi,
                SUM(d.unit_count) as miktar, COUNT(DISTINCT d.tarih) as gun_sayisi
            FROM daily_metrics_summary d
            JOIN magazalar m ON d.magaza_id = m.id
            WHERE {where_clause}
            GROUP BY m.ad, {be}
            ORDER BY ciro DESC
        """
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # Excel oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mağaza Analizi"
        
        # Başlıklar
        headers = ["Mağaza Adı", "Bölge", "Ciro (₺)", "Fiş Adedi", "Miktar", "Sepet Ortalaması", "Günlük Ort. Ciro"]
        ws.append(headers)
        
        # Stil
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Verileri ekle
        for r in rows:
            ciro = float(r['ciro'] or 0)
            fis = int(r['fis_adedi'] or 0)
            ws.append([
                r['magaza_ad'],
                r['bolge'],
                ciro,
                fis,
                int(r['miktar'] or 0),
                round(ciro / fis, 2) if fis > 0 else 0,
                round(ciro / r['gun_sayisi'], 2) if r['gun_sayisi'] > 0 else 0
            ])
            
        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2
            
        # Response hazırla
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Magaza_Analizi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return HttpResponse(f"Hata oluştu: {str(e)}", status=500)
    finally:
        db_engine.release_connection(conn)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_store_list(request):
    """Mağaza listesini döndürür (id, ad, bolge) - Beklenen müşteri filtresi için."""
    conn = None
    try:
        conn = db_engine.get_connection()
        cursor = db_engine.get_dict_cursor(conn)
        cursor.execute("""
            SELECT id, ad, bolge 
            FROM magazalar 
            WHERE ad IS NOT NULL 
            ORDER BY ad
        """)
        stores = [dict(r) for r in cursor.fetchall()]
        return Response({'stores': stores})
    except Exception as e:
        logger.error(f"Store list error: {e}")
        return Response({'error': str(e)}, status=500)
    finally:
        if conn: db_engine.release_connection(conn)
